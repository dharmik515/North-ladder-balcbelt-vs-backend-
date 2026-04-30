import argparse
import csv
import os
import re
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Monkey patch to suppress corrupt validation rules
original_init = DataValidation.__init__

def patched_init(self, *args, **kwargs):
    try:
        original_init(self, *args, **kwargs)
    except ValueError:
        # Skip corrupted validation rules
        pass

DataValidation.__init__ = patched_init

try:
    from rapidfuzz import fuzz
except ImportError:
    fuzz = None


def normalize_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return text.strip()


def normalize_imei(value: Any) -> Optional[str]:
    if pd.isna(value) or value == "":
        return None
    text = str(value).strip()
    text = re.sub(r"[^0-9A-Za-z]", "", text)
    if text == "":
        return None
    return text


def normalize_storage(value: Any) -> Optional[str]:
    if pd.isna(value) or value == "":
        return None
    text = str(value).lower()
    text = re.sub(r"[^0-9gmb]+", "", text)
    text = re.sub(r"gb", "", text)
    text = text.strip()
    return text if text else None


def similarity_score(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if fuzz:
        return fuzz.token_sort_ratio(a, b)
    return SequenceMatcher(None, a, b).ratio() * 100


def load_worksheet_as_frame(path: str, sheet_name: Optional[str] = None, max_rows: Optional[int] = None) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=sheet_name, nrows=max_rows)
    except Exception as e:
        print(f"Warning: Standard read failed ({e}). Attempting openpyxl fallback...")
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            rows = []
            header = None
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i == 1:
                    header = [str(cell) if cell is not None else "" for cell in row]
                    continue
                rows.append([cell for cell in row])
                if max_rows and len(rows) >= max_rows:
                    break
            if header is None:
                return pd.DataFrame()
            return pd.DataFrame(rows, columns=header)
        except Exception as e2:
            print(f"Fallback also failed: {e2}")
            print("Attempting to read with xlrd...")
            return pd.read_excel(path, sheet_name=sheet_name, engine='openpyxl', nrows=max_rows)


@dataclass
class Record:
    source: str
    row_index: int
    imei: Optional[str]
    imei2: Optional[str]
    brand: str
    model: str
    model_number: str
    storage: str
    color: str
    serial: str
    asset_label: str
    raw: Dict[str, Any]


def build_blackbelt_records(path: str) -> List[Record]:
    df = load_worksheet_as_frame(path, sheet_name="Sheet1")
    records: List[Record] = []
    for idx, row in df.iterrows():
        imei = normalize_imei(row.get("IMEI/MEID", row.get("IMEI", row.get("IMEI1", None))))
        imei2 = normalize_imei(row.get("IMEI2", None))
        brand = normalize_text(row.get("Manufacturer", ""))
        model = normalize_text(row.get("Model", row.get("Model Number", "")))
        model_number = normalize_text(row.get("Model Number", ""))
        storage = normalize_storage(row.get("Handset Memory Size", row.get("Memory", "")))
        color = normalize_text(row.get("Device Colour", row.get("Color", "")))
        serial = normalize_text(row.get("Serial Number", ""))
        asset_label = normalize_text(row.get("Model", ""))

        records.append(
            Record(
                source="blackbelt",
                row_index=idx,
                imei=imei,
                imei2=imei2,
                brand=brand,
                model=model,
                model_number=model_number,
                storage=storage,
                color=color,
                serial=serial,
                asset_label=asset_label,
                raw=row.to_dict(),
            )
        )
    return records


def build_company_records(path: str) -> List[Record]:
    df = load_worksheet_as_frame(path, sheet_name="BulkSell")
    records: List[Record] = []
    for idx, row in df.iterrows():
        imei = normalize_imei(row.get("IMEI Number", row.get("IMEI", None)))
        brand = normalize_text(row.get("Brand", row.get("Asset Label", "")))
        asset_label = normalize_text(row.get("Asset Label", ""))
        model = normalize_text(row.get("Asset Label", row.get("Category", "")))
        model_number = normalize_text(row.get("Brand", ""))
        storage = normalize_storage(asset_label)
        color = normalize_text(row.get("Asset Label", ""))
        serial = normalize_text(row.get("Barcode", row.get("QR Code", "")))

        records.append(
            Record(
                source="company",
                row_index=idx,
                imei=imei,
                imei2=None,
                brand=brand,
                model=model,
                model_number=model_number,
                storage=storage,
                color=color,
                serial=serial,
                asset_label=asset_label,
                raw=row.to_dict(),
            )
        )
    return records


def compute_match_score(company: Record, blackbelt: Record, match_type: str) -> float:
    """Compute a detailed match score with weighted attributes."""
    if match_type == "exact_imei":
        return 100.0
    if match_type == "exact_imei2":
        return 100.0

    score = 0.0
    weights = {
        "brand": 0.25,
        "model": 0.35,
        "storage": 0.15,
        "color": 0.10,
        "serial": 0.15,
    }

    # Brand match
    brand_sim = similarity_score(company.brand, blackbelt.brand)
    score += brand_sim * weights["brand"]

    # Model match
    model_sim = similarity_score(company.model, blackbelt.model)
    score += model_sim * weights["model"]

    # Storage match
    if company.storage and blackbelt.storage:
        storage_match = 100.0 if company.storage == blackbelt.storage else 0.0
    else:
        storage_match = 50.0 if (company.storage or blackbelt.storage) else 100.0
    score += storage_match * weights["storage"]

    # Color match
    if company.color and blackbelt.color:
        color_match = 100.0 if company.color == blackbelt.color else similarity_score(company.color, blackbelt.color)
    else:
        color_match = 50.0 if (company.color or blackbelt.color) else 100.0
    score += color_match * weights["color"]

    # Serial/identifier match
    if company.serial and blackbelt.serial:
        serial_match = 100.0 if company.serial == blackbelt.serial else (50.0 if company.serial in blackbelt.serial or blackbelt.serial in company.serial else 0.0)
    else:
        serial_match = 50.0 if (company.serial or blackbelt.serial) else 100.0
    score += serial_match * weights["serial"]

    return score


def find_matches(company: Record, blackbelt_index: Dict[str, List[Record]], blackbelt_records: List[Record], limit: int = 10000) -> List[Tuple[Record, float, str, str]]:
    matches: List[Tuple[Record, float, str, str]] = []

    # Layer 1: Exact IMEI match
    if company.imei:
        exact_hits = blackbelt_index.get(company.imei, [])
        for hit in exact_hits:
            matches.append((hit, 100.0, "exact_imei", "Exact IMEI match"))
            return matches  # Return immediately on exact match

    # Layer 2: Alternate IMEI (IMEI2) match
    if company.imei:
        for blackbelt_record in blackbelt_records[:limit]:
            if blackbelt_record.imei2 == company.imei:
                matches.append((blackbelt_record, 100.0, "exact_imei2", "Matched via IMEI2 (alternate IMEI scanned)"))
                return matches

    # Layer 3: Fuzzy model matching and attribute similarity
    # Only compare against records with same brand or model keywords
    candidates = []
    for blackbelt_record in blackbelt_records[:limit]:
        # Skip if already exactly matched
        if blackbelt_record.imei == company.imei or blackbelt_record.imei2 == company.imei:
            continue
        
        # Filter by brand similarity (fast)
        if company.brand and blackbelt_record.brand:
            if similarity_score(company.brand, blackbelt_record.brand) < 40:
                continue
        
        # Filter by storage equality (fast)
        if company.storage and blackbelt_record.storage:
            if company.storage != blackbelt_record.storage:
                continue
        
        candidates.append(blackbelt_record)
    
    # Now do more expensive fuzzy scoring on filtered candidates
    for blackbelt_record in candidates:
        score = compute_match_score(company, blackbelt_record, "fuzzy")
        if score >= 60.0:
            reason = f"Fuzzy match (score: {score:.1f}%)"
            matches.append((blackbelt_record, score, "fuzzy_model", reason))

    matches.sort(key=lambda item: item[1], reverse=True)
    return matches


def build_index(records: List[Record]) -> Dict[str, List[Record]]:
    index: Dict[str, List[Record]] = defaultdict(list)
    for record in records:
        if record.imei:
            index[record.imei].append(record)
        if record.imei2:
            index[record.imei2].append(record)
    return index


def generate_reports(company_records: List[Record], blackbelt_records: List[Record], output_dir: str):
    os.makedirs(output_dir, exist_ok=True)
    blackbelt_index = build_index(blackbelt_records)

    high_confidence_rows = []
    medium_confidence_rows = []
    low_confidence_rows = []
    unmatched_rows = []

    for company in company_records:
        matches = find_matches(company, blackbelt_index, blackbelt_records)
        if not matches:
            unmatched_rows.append(build_report_row(company, None, 0.0, "no_match", "No matching record found"))
            continue

        top_match, score, reason, description = matches[0]
        
        if score >= 90.0:
            high_confidence_rows.append(build_report_row(company, top_match, score, reason, description))
        elif score >= 70.0:
            medium_confidence_rows.append(build_report_row(company, top_match, score, reason, description))
        else:
            low_confidence_rows.append(build_report_row(company, top_match, score, reason, description))

    write_rows(os.path.join(output_dir, "high_confidence_matches.csv"), high_confidence_rows)
    write_rows(os.path.join(output_dir, "medium_confidence_matches.csv"), medium_confidence_rows)
    write_rows(os.path.join(output_dir, "low_confidence_matches.csv"), low_confidence_rows)
    write_rows(os.path.join(output_dir, "unmatched.csv"), unmatched_rows)

    total = len(high_confidence_rows) + len(medium_confidence_rows) + len(low_confidence_rows) + len(unmatched_rows)
    print(f"\n=== MATCHING RESULTS ===")
    print(f"High confidence matches: {len(high_confidence_rows)} ({100*len(high_confidence_rows)/total:.1f}%)")
    print(f"Medium confidence matches: {len(medium_confidence_rows)} ({100*len(medium_confidence_rows)/total:.1f}%)")
    print(f"Low confidence matches: {len(low_confidence_rows)} ({100*len(low_confidence_rows)/total:.1f}%)")
    print(f"Unmatched: {len(unmatched_rows)} ({100*len(unmatched_rows)/total:.1f}%)")


def build_report_row(company: Record, blackbelt: Optional[Record], score: float, reason: str, description: str) -> Dict[str, Any]:
    recommendation = "AUTO_CORRECT" if score >= 95.0 else ("REVIEW" if score >= 70.0 else "MANUAL_REVIEW")
    
    row = {
        "decision": recommendation,
        "confidence_score": round(score, 2),
        "match_reason": reason,
        "description": description,
        "company_row_index": company.row_index,
        "company_imei": company.imei or "N/A",
        "company_brand": company.brand,
        "company_model": company.model,
        "company_storage": company.storage or "N/A",
        "company_color": company.color or "N/A",
        "company_asset_label": company.asset_label,
        "company_serial": company.serial or "N/A",
    }
    if blackbelt:
        correction_needed = (
            (company.imei != blackbelt.imei) and 
            (company.brand != blackbelt.brand or company.model != blackbelt.model)
        )
        row.update(
            {
                "correction_needed": "YES" if correction_needed else "NO",
                "blackbelt_row_index": blackbelt.row_index,
                "blackbelt_imei": blackbelt.imei or "N/A",
                "blackbelt_imei2": blackbelt.imei2 or "N/A",
                "blackbelt_brand": blackbelt.brand,
                "blackbelt_model": blackbelt.model,
                "blackbelt_storage": blackbelt.storage or "N/A",
                "blackbelt_color": blackbelt.color or "N/A",
                "blackbelt_asset_label": blackbelt.asset_label,
                "blackbelt_serial": blackbelt.serial or "N/A",
                "suggested_correction": f"Update company IMEI from {company.imei} to {blackbelt.imei}" if (company.imei != blackbelt.imei) else (
                    f"Update model from {company.model} to {blackbelt.model}" if (company.model != blackbelt.model) else ""
                ),
            }
        )
    else:
        row.update(
            {
                "correction_needed": "N/A",
                "blackbelt_row_index": None,
                "blackbelt_imei": "N/A",
                "blackbelt_imei2": "N/A",
                "blackbelt_brand": "N/A",
                "blackbelt_model": "N/A",
                "blackbelt_storage": "N/A",
                "blackbelt_color": "N/A",
                "blackbelt_asset_label": "N/A",
                "blackbelt_serial": "N/A",
                "suggested_correction": "Manual research required",
            }
        )
    return row


def write_rows(path: str, rows: List[Dict[str, Any]]):
    if not rows:
        return
    with open(path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Blackbelt-NorthLadder mismatch detection pipeline")
    parser.add_argument("--blackbelt", required=True, help="Blackbelt Excel file path")
    parser.add_argument("--company", required=True, help="NorthLadder company Excel file path")
    parser.add_argument("--output", default="output", help="Output directory for report CSV files")
    return parser.parse_args()


def main():
    args = parse_args()
    blackbelt_records = build_blackbelt_records(args.blackbelt)
    company_records = build_company_records(args.company)
    generate_reports(company_records, blackbelt_records, args.output)


if __name__ == "__main__":
    main()
