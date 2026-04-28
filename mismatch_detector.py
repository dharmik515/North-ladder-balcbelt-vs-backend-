"""
Mismatch detector for company (Stack) data vs. Blackbelt reference.

Flags rows in the company file where data was likely entered wrong. Runs
five independent layers — each produces its own flags, sorted by severity:

  L1 FORMAT      — IMEI Luhn check, length, character class.
  L2 SCAN-SLOT   — wrong type of value in the wrong column (e.g. serial
                   scanned into the IMEI slot, IMEI scanned into Barcode).
  L3 INTRA-ROW   — Brand / Asset Label / Category disagree with each other.
  L4 CATALOG     — CO's (brand, model, storage) triple doesn't exist in
                   Blackbelt's catalog of real devices.
  L5 DUPLICATES  — same IMEI or AssetId appears multiple times in CO.

The two files don't need to share any device IDs for this to work — BB
is used as a REFERENCE CATALOG of valid specs, not as a row-level
match target.

Output:
  results_new/flagged.csv      — every flag, sorted by severity
  results_new/summary.json     — counts per layer / severity / issue
  results_new/per_row.csv      — one row per company record with its worst flag
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation

# silence corrupt validation rules in the workbook
_orig = DataValidation.__init__
def _patched(self, *a, **kw):
    try: _orig(self, *a, **kw)
    except ValueError: pass
DataValidation.__init__ = _patched


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

BB_PATH = r"C:\Users\dharm\Downloads\ExcelReports-analyst-14-04-2026-12-12-18.xlsx"
CO_PATH = r"C:\Users\dharm\Downloads\Stack Bulk Upload - 2026-04-14T153918.672.xlsx"
OUT_DIR = Path("results_new")

SEVERITY_RANK = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}


# ---------------------------------------------------------------------------
# Primitives
# ---------------------------------------------------------------------------

def norm_text(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return re.sub(r"\s+", " ", str(v)).strip().lower()


def clean_id(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)) or v == "":
        return ""
    return re.sub(r"[^0-9A-Za-z]", "", str(v)).upper()


def luhn_valid(s: str) -> bool:
    """Standard Luhn check for IMEIs (15 numeric digits)."""
    if not s.isdigit() or len(s) != 15:
        return False
    total = 0
    for i, ch in enumerate(reversed(s)):
        n = int(ch)
        if i % 2 == 1:
            n *= 2
            if n > 9:
                n -= 9
        total += n
    return total % 10 == 0


def imei_shape(s: str) -> str:
    """Classify what a cleaned string *looks* like."""
    if not s:
        return "empty"
    if s.isdigit():
        if len(s) == 15:
            return "imei15"
        if 14 <= len(s) <= 16:
            return "imei_like"
        if 8 <= len(s) <= 13:
            return "short_numeric"
        return "numeric_other"
    # alphanumeric
    if re.fullmatch(r"[A-Z]{1,3}[0-9A-Z]{7,13}", s):
        return "serial_like"
    return "alnum_other"


def extract_storage_gb(text: str) -> Optional[int]:
    """Pull first storage spec out of free text. Prefer TB/GB over SSD-less digits."""
    if not text:
        return None
    t = str(text).upper()
    m = re.search(r"(\d+)\s*(TB|GB)\b", t)
    if not m:
        return None
    n = int(m.group(1))
    return n * 1024 if m.group(2) == "TB" else n


def brand_canonical(s: str) -> str:
    s = norm_text(s)
    aliases = {
        "apple inc": "apple", "samsung electronics": "samsung",
        "samsung korea": "samsung",
        "google inc": "google", "google llc": "google",
        "xiaomi communications co ltd": "xiaomi", "xiaomi corp": "xiaomi",
        "microsoft surface": "microsoft",
        "macbooks": "apple",
    }
    return aliases.get(s, s)


# Brand fingerprints — used to detect Brand↔Label mismatch.
# Each brand lists tokens that, if found in the Asset Label, confirm the brand.
BRAND_TOKENS = {
    "apple":     {"apple", "iphone", "ipad", "macbook", "airpod", "watch", "imac"},
    "samsung":   {"samsung", "galaxy"},
    "google":    {"google", "pixel"},
    "xiaomi":    {"xiaomi", "redmi", "poco", "mi "},
    "huawei":    {"huawei", "mate", "pura", "nova", "matepad"},
    "oppo":      {"oppo", "reno", "find x"},
    "oneplus":   {"oneplus"},
    "honor":     {"honor"},
    "hp":        {"hp ", "elite", "pavilion", "envy", "omen", "probook", "zbook"},
    "dell":      {"dell", "latitude", "inspiron", "xps", "vostro", "precision"},
    "lenovo":    {"lenovo", "thinkpad", "yoga", "ideapad"},
    "asus":      {"asus", "vivobook", "zenbook", "rog"},
    "microsoft": {"surface"},
    "nothing":   {"nothing"},
    "vivo":      {"vivo"},
    "garmin":    {"garmin", "fenix", "forerunner", "vivoactive"},
    "lg":        {"lg ", "lg,"},
}

# Expected category given a model family token in the label
CATEGORY_RULES = [
    ({"iphone"}, {"mobile phone"}),
    ({"ipad"}, {"tablet"}),
    ({"macbook", "latitude", "thinkpad", "elitebook", "zbook", "vivobook", "zenbook",
      "pavilion", "inspiron", "xps", "yoga", "ideapad"}, {"laptop"}),
    ({"airpod"}, {"earbuds", "audio", "accessory", "others"}),
    ({"watch", "fenix", "forerunner"}, {"watch", "wearable", "smartwatch", "others"}),
]


# ---------------------------------------------------------------------------
# Load + parse
# ---------------------------------------------------------------------------

def load_blackbelt(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Sheet1")
    bb = pd.DataFrame({
        "bb_row":        df.index,
        "imei":          df["IMEI/MEID"].map(clean_id),
        "imei2":         df["IMEI2"].map(clean_id),
        "serial":        df["Serial Number"].map(clean_id),
        "brand":         df["Manufacturer"].map(brand_canonical),
        "model":         df["Model"].map(norm_text),
        "model_number":  df["Model Number"].map(norm_text),
        "storage_gb":    df["Handset Memory Size"].map(extract_storage_gb),
        "color":         df["Device Colour"].map(norm_text),
    })
    return bb


# Column fingerprints — column names that uniquely identify each format.
# We score every sheet against both; the best match wins. Sheet names and
# file names are NOT used, so a renamed workbook still loads correctly.
_STACKBULK_COLS = {"AssetId", "IMEI Number", "Appraisal", "Asset Label",
                   "Latest Assessed Grade", "Barcode"}
_MASTER_COLS    = {"IMEI", "Deal Id", "Stack", "Room", "Bin", "Location",
                   "Brand", "Model", "Grade"}


def _detect_company_format(path: str) -> tuple[str, str]:
    """
    Return (format_name, sheet_name) by inspecting column headers across all
    sheets. format_name is one of {"stackbulk", "master"}.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        best = (0, None, None)  # (score, fmt, sheet)
        for sn in wb.sheetnames:
            ws = wb[sn]
            header = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                header = [str(c).strip() for c in row if c is not None]
                break
            if not header:
                continue
            cols = set(header)
            stack_score  = len(cols & _STACKBULK_COLS)
            master_score = len(cols & _MASTER_COLS)
            if stack_score >= 4 and stack_score > best[0]:
                best = (stack_score, "stackbulk", sn)
            if master_score >= 4 and master_score > best[0]:
                best = (master_score, "master", sn)
    finally:
        wb.close()

    if best[1] is None:
        raise ValueError(
            "Unrecognised company file: no sheet matched the Stack Bulk "
            "(AssetId/IMEI Number/Appraisal/Asset Label) or Master Template "
            "(IMEI/Deal Id/Stack/Room/Bin) column signature."
        )
    return best[1], best[2]


def load_company(path: str) -> pd.DataFrame:
    """
    Load a company-side workbook. Auto-detects the format by column headers
    (not by sheet name or filename, both of which can change between exports):
      - Stack Bulk Upload    — has AssetId / IMEI Number / Appraisal / Asset Label
      - Master Template      — has IMEI / Deal Id / Stack / Room / Bin

    Both formats are normalised into the same internal schema so the rest of
    the pipeline is format-agnostic.
    """
    fmt, sheet = _detect_company_format(path)
    if fmt == "stackbulk":
        return _load_company_stackbulk(path, sheet)
    return _load_company_master(path, sheet)


def _load_company_stackbulk(path: str, sheet: str = "BulkSell") -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet)
    # Stack Bulk has no warehouse Room/Bin/Location — fall back to the
    # storage-member fields so the Location column still shows something.
    loc_parts = []
    for col in ("Storage Member HO", "Storage Member Country"):
        if col in df.columns:
            loc_parts.append(df[col].astype(str).where(df[col].notna(), ""))
    location_text = (loc_parts[0] if loc_parts else pd.Series([""] * len(df)))
    for p in loc_parts[1:]:
        location_text = location_text.str.cat(p, sep=" / ", na_rep="")

    # VAT can come from either of two columns depending on the export
    # variant; prefer the appraisal-side value, fall back to the sell-side.
    vat = df.get("Appraisal VATType")
    if vat is None or vat.isna().all():
        vat = df.get("VAT Type", pd.Series([""] * len(df)))
    vat = vat.map(norm_text).fillna("")

    storage_country = df.get("Storage Member Country", pd.Series([""] * len(df)))
    storage_country = storage_country.map(norm_text).fillna("")

    co = pd.DataFrame({
        "co_row":       df.index,
        "appraisal":    df["Appraisal"].astype(str),
        "asset_id":     df["AssetId"].astype(str),
        "imei_raw":     df["IMEI Number"].astype(str).where(df["IMEI Number"].notna(), ""),
        "imei":         df["IMEI Number"].map(clean_id),
        "barcode_raw":  df["Barcode"].astype(str).where(df["Barcode"].notna(), ""),
        "barcode":      df["Barcode"].map(clean_id),
        "qr":           df["QR Code"].astype(str).where(df["QR Code"].notna(), ""),
        "brand":        df["Brand"].map(brand_canonical),
        "asset_label":  df["Asset Label"].map(norm_text),
        "category":     df["Category"].map(norm_text),
        "grade":        df["Latest Assessed Grade"].map(norm_text),
        "vat_type":     vat,
        "country":      storage_country,
        "location_text": location_text,
    })
    co["storage_gb"] = co["asset_label"].map(extract_storage_gb)
    co["imei_shape"] = co["imei"].map(imei_shape)
    co["barcode_shape"] = co["barcode"].map(imei_shape)
    co["deal_date"] = co["appraisal"].map(_parse_deal_date)
    return co


def _parse_deal_date(s) -> Optional[str]:
    """
    Master Deal IDs (and Stack Appraisals) are formatted like 'AE-DDMMYY-XXXXXX'
    or 'RENAE-DDMMYY-XXXXXX'. Return ISO date string YYYY-MM-DD, or None.
    """
    if not s or (isinstance(s, float) and pd.isna(s)):
        return None
    m = re.search(r"(\d{2})(\d{2})(\d{2})-\d+", str(s))
    if not m:
        return None
    dd, mm, yy = m.groups()
    try:
        # Two-digit year: assume 20YY (current era).
        from datetime import date
        return date(2000 + int(yy), int(mm), int(dd)).isoformat()
    except ValueError:
        return None


def _load_company_master(path: str, sheet: str = "StockTake Template") -> pd.DataFrame:
    """
    Master Template loader.

    The Master/StockTake export has fewer columns than Stack Bulk — there is
    no Barcode, QR Code, AssetId, or descriptive Asset Label. We map what we
    can and leave the rest blank; layers that consume the missing fields
    (L2 SCAN-SLOT, L6 IMEI1/IMEI2, L15 QR-vs-IMEI) simply produce no flags
    on Master input, which is fine.

    Empty bin rows (no IMEI and no Brand) are dropped — they're warehouse
    placeholders, not records to validate.
    """
    df = pd.read_excel(path, sheet_name=sheet)

    # Drop rows that look like empty bin slots (no IMEI AND no Brand).
    has_imei  = df["IMEI"].notna() & (df["IMEI"].astype(str).str.strip() != "")
    has_brand = df["Brand"].notna() & (df["Brand"].astype(str).str.strip() != "")
    df = df[has_imei | has_brand].reset_index(drop=True)

    # Combine warehouse Room / Bin / Location into one human-readable string.
    def _join_location(row):
        parts = [str(row.get(c, "") or "").strip() for c in ("Room", "Bin", "Location")]
        return " / ".join(p for p in parts if p)
    location_text = df.apply(_join_location, axis=1)

    co = pd.DataFrame({
        "co_row":       df.index,
        "appraisal":    df["Deal Id"].astype(str).where(df["Deal Id"].notna(), ""),
        "asset_id":     df["Deal Id"].astype(str).where(df["Deal Id"].notna(), ""),
        "imei_raw":     df["IMEI"].astype(str).where(df["IMEI"].notna(), ""),
        "imei":         df["IMEI"].map(clean_id),
        "barcode_raw":  "",
        "barcode":      "",
        "qr":           "",
        "brand":        df["Brand"].map(brand_canonical),
        "asset_label":  df["Model"].map(norm_text),
        "category":     df["Category"].map(norm_text),
        "grade":        df["Grade"].map(norm_text),
        "vat_type":     df.get("VAT Type", pd.Series([""] * len(df))).map(norm_text).fillna(""),
        "country":      df.get("Country",  pd.Series([""] * len(df))).map(norm_text).fillna(""),
        "location_text": location_text,
    })
    co["storage_gb"] = co["asset_label"].map(extract_storage_gb)
    co["imei_shape"] = co["imei"].map(imei_shape)
    co["barcode_shape"] = co["barcode"].map(imei_shape)
    co["deal_date"] = co["appraisal"].map(_parse_deal_date)
    return co


def build_catalog(bb: pd.DataFrame) -> dict:
    """
    Catalog keyed by (brand, model_family).
    model_family = bb.model with storage/trailing-variant tokens removed.
    """
    catalog: dict = defaultdict(lambda: {
        "storages": set(), "model_numbers": set(), "colors": set(), "raw_models": set()
    })
    for _, r in bb.iterrows():
        if not r["brand"] or not r["model"]:
            continue
        key = (r["brand"], r["model"])
        if r["storage_gb"] and not pd.isna(r["storage_gb"]):
            catalog[key]["storages"].add(int(r["storage_gb"]))
        if r["model_number"]:
            catalog[key]["model_numbers"].add(r["model_number"])
        if r["color"]:
            catalog[key]["colors"].add(r["color"])
        catalog[key]["raw_models"].add(r["model"])
    return dict(catalog)


# ---------------------------------------------------------------------------
# Flag objects
# ---------------------------------------------------------------------------

@dataclass
class Flag:
    co_row: int
    asset_id: str
    appraisal: str
    layer: str
    issue: str
    severity: str  # CRITICAL / HIGH / MEDIUM / LOW
    field: str
    current_value: str
    expected: str
    suggested_fix: str


def mk(co: pd.Series, layer: str, issue: str, sev: str, field: str,
       cur: str, exp: str, fix: str) -> Flag:
    return Flag(
        co_row=int(co["co_row"]),
        asset_id=str(co["asset_id"]),
        appraisal=str(co["appraisal"]),
        layer=layer, issue=issue, severity=sev, field=field,
        current_value=str(cur), expected=str(exp), suggested_fix=fix,
    )


# ---------------------------------------------------------------------------
# L1: Format integrity
# ---------------------------------------------------------------------------

def layer1_format(co: pd.DataFrame) -> list[Flag]:
    flags: list[Flag] = []
    for _, r in co.iterrows():
        imei = r["imei"]
        if r["category"] == "mobile phone":
            if not imei:
                flags.append(mk(r, "L1", "imei_missing", "HIGH", "IMEI Number",
                                "", "15-digit numeric IMEI",
                                "Re-scan IMEI from device (*#06#)."))
            elif r["imei_shape"] == "imei15":
                if not luhn_valid(imei):
                    flags.append(mk(r, "L1", "imei_luhn_fail", "HIGH", "IMEI Number",
                                    imei, "IMEI that passes Luhn checksum",
                                    "Likely a typo or mis-read digit; re-scan."))
            elif r["imei_shape"] in ("short_numeric", "numeric_other"):
                flags.append(mk(r, "L1", "imei_wrong_length", "HIGH", "IMEI Number",
                                imei, "15 digits",
                                f"IMEI has {len(imei)} digits; re-scan."))
            elif r["imei_shape"] == "imei_like":
                # 14-digit numerics are just short IMEIs (or MEIDs in disguise).
                # 16-digit numerics may be IMEISV (IMEI + 2-digit Software
                # Version). Detect that case so we don't flag a legitimate
                # IMEISV as a hard error.
                if len(imei) == 16:
                    body14 = imei[:14]
                    matched_imei = next(
                        (body14 + str(x) for x in range(10) if luhn_valid(body14 + str(x))),
                        None,
                    )
                    if matched_imei:
                        flags.append(mk(
                            r, "L1", "looks_like_imeisv", "LOW", "IMEI Number",
                            imei,
                            f"15-digit IMEI; IMEISV's first 14 digits give {matched_imei}",
                            "IMEISV is IMEI + 2-digit Software Version. The IMEI portion is "
                            "Luhn-valid; replace this 16-digit value with the 15-digit IMEI "
                            "for consistency, or leave it if your downstream system tracks "
                            "IMEISV intentionally.",
                        ))
                    else:
                        flags.append(mk(
                            r, "L1", "imei_wrong_length", "HIGH", "IMEI Number",
                            imei, "15 digits",
                            "16 digits but not a valid IMEISV (first 14 digits are not a "
                            "Luhn-valid IMEI body). Re-scan from the device.",
                        ))
                else:  # 14 digits
                    flags.append(mk(
                        r, "L1", "imei_wrong_length", "HIGH", "IMEI Number",
                        imei, "15 digits",
                        f"IMEI has {len(imei)} digits; re-scan.",
                    ))
            # serial_like / alnum_other on a mobile row is handled by L2 scan-slot
    return flags


# ---------------------------------------------------------------------------
# L2: Scan-slot confusion
# ---------------------------------------------------------------------------

def layer2_scan_slot(co: pd.DataFrame) -> list[Flag]:
    flags: list[Flag] = []
    for _, r in co.iterrows():
        # IMEI column holding an alphanumeric value on a phone or tablet row.
        # The previous false-positive concern is addressed by the category
        # filter at run() — laptops/smartwatches/earphones never reach here.
        if (r["category"] in ("mobile phone", "tablet")
                and r["imei_shape"] in ("serial_like", "alnum_other")):
            flags.append(mk(r, "L2", "serial_in_imei_slot", "HIGH",
                            "IMEI Number", r["imei"], "15-digit numeric IMEI",
                            f"Looks like a serial number ('{r['imei']}') was scanned "
                            f"into the IMEI slot. Re-scan the IMEI from the device "
                            f"(dial *#06# on phones)."))

        # Barcode column holding a 15-digit number (IMEI) on a phone row
        if r["category"] == "mobile phone" and r["barcode_shape"] == "imei15":
            if luhn_valid(r["barcode"]):
                flags.append(mk(r, "L2", "imei_in_barcode_slot", "HIGH",
                                "Barcode", r["barcode"], "internal barcode",
                                f"Value in Barcode ('{r['barcode']}') looks like a "
                                f"valid IMEI. Check whether IMEI was scanned into the "
                                f"wrong column."))

        # Barcode AND IMEI both hold the same value = almost certainly one scan pasted twice
        if r["imei"] and r["imei"] == r["barcode"]:
            flags.append(mk(r, "L2", "imei_equals_barcode", "MEDIUM",
                            "Barcode/IMEI Number", r["imei"], "distinct values",
                            "IMEI and Barcode hold identical value; one slot was "
                            "likely filled by mistake."))
    return flags


# ---------------------------------------------------------------------------
# L3: Intra-row consistency
# ---------------------------------------------------------------------------

def layer3_intra_row(co: pd.DataFrame) -> list[Flag]:
    flags: list[Flag] = []
    for _, r in co.iterrows():
        brand, label, cat = r["brand"], r["asset_label"], r["category"]

        # 3a. Brand column disagrees with Asset Label
        if brand and label and brand in BRAND_TOKENS:
            tokens = BRAND_TOKENS[brand]
            if not any(t in label for t in tokens):
                # check if a DIFFERENT brand's tokens are present
                claimed_other = [
                    b for b, toks in BRAND_TOKENS.items()
                    if b != brand and any(t in label for t in toks)
                ]
                if claimed_other:
                    flags.append(mk(r, "L3", "brand_label_contradiction", "HIGH",
                                    "Brand",
                                    f"Brand='{brand}', Label='{label}'",
                                    f"Brand should be one of {claimed_other}",
                                    f"Asset Label looks like {claimed_other[0]}; "
                                    f"either Brand is wrong or Label is wrong."))
                else:
                    flags.append(mk(r, "L3", "brand_token_absent", "MEDIUM",
                                    "Brand/Asset Label",
                                    f"Brand='{brand}', Label='{label}'",
                                    f"Label should mention one of {sorted(tokens)}",
                                    "Verify this listing describes the claimed brand."))

        # 3b. Category disagrees with model family in the label
        if label and cat:
            for family_tokens, valid_cats in CATEGORY_RULES:
                if any(t in label for t in family_tokens):
                    if cat not in valid_cats:
                        flags.append(mk(r, "L3", "category_model_mismatch", "HIGH",
                                        "Category",
                                        f"Category='{cat}', Label='{label}'",
                                        f"One of {sorted(valid_cats)}",
                                        f"Label mentions {list(family_tokens)[0]} "
                                        f"→ category should be {sorted(valid_cats)}."))
                    break

        # 3c. Storage missing from label for a device type that always has storage
        if cat in ("mobile phone", "tablet", "laptop") and r["storage_gb"] is None and label:
            flags.append(mk(r, "L3", "storage_missing", "MEDIUM",
                            "Asset Label", label, "label including e.g. '128GB'",
                            "Storage capacity not found in Asset Label — either "
                            "template not filled or wrong label."))
    return flags


# ---------------------------------------------------------------------------
# L4: Catalog validation
# ---------------------------------------------------------------------------

def layer4_catalog(co: pd.DataFrame, brand_idx: dict) -> list[Flag]:
    """
    Validate CO (brand, model, storage) against the Blackbelt catalog of
    (brand, model) → {storages, model_numbers, colors}. Uses the shared
    brand_idx so the expensive per-brand token indexing happens only once.
    """
    flags: list[Flag] = []
    for _, r in co.iterrows():
        bb_match = best_bb_model(r["brand"], r["asset_label"], brand_idx)
        if not bb_match:
            continue
        model_str, _, data = bb_match
        if r["storage_gb"] and data["storages"]:
            if int(r["storage_gb"]) not in data["storages"]:
                flags.append(mk(r, "L4", "storage_unseen_in_bb", "MEDIUM",
                                "Asset Label (storage)",
                                f"{int(r['storage_gb'])}GB for {r['brand']} {model_str}",
                                f"one of {sorted(data['storages'])}GB (BB-seen)",
                                f"Blackbelt has only tested {r['brand']} {model_str} in "
                                f"{sorted(data['storages'])}GB. {int(r['storage_gb'])}GB "
                                f"may be legit (BB coverage is partial) but verify the label."))
    return flags


# ---------------------------------------------------------------------------
# L6: IMEI1 ↔ IMEI2 confusion (the error the user specifically described)
# ---------------------------------------------------------------------------

def layer6_imei1_vs_imei2(co: pd.DataFrame) -> list[Flag]:
    """
    When a device emits both IMEIs (dual-SIM, most modern phones), IMEI1 and
    IMEI2 are usually consecutive or near-consecutive integers (same TAC+FAC,
    adjacent serial). If the operator was told to scan IMEI1 but scanned
    IMEI2 by mistake, we won't necessarily catch it alone — but when two
    company rows have IMEIs that are consecutive (|diff| <= 5) AND belong
    to the SAME AssetId (same model), that's a strong signal that two
    physical devices were scanned consistently OR that IMEI1/IMEI2 of a
    single device were saved as two rows by mistake.
    """
    flags: list[Flag] = []
    numeric = co[(co["imei_shape"] == "imei15")].copy()
    if len(numeric) < 2:
        return flags
    # Python int handles 15-digit numbers fine; numpy int64 overflows.
    numeric["imei_int"] = numeric["imei"].map(lambda s: int(s))

    # group by AssetId then look for close neighbours
    for aid, group in numeric.groupby("asset_id"):
        if len(group) < 2:
            continue
        ordered = group.sort_values("imei", key=lambda s: s.map(int))
        vals = [int(x) for x in ordered["imei"].tolist()]
        rows = ordered.reset_index(drop=True)
        for i in range(len(vals) - 1):
            diff = vals[i + 1] - vals[i]
            if 1 <= diff <= 3:
                r1, r2 = rows.iloc[i], rows.iloc[i + 1]
                flags.append(mk(r2, "L6", "possible_imei1_imei2_pair", "MEDIUM",
                                "IMEI Number",
                                f"{r1['imei']} / {r2['imei']} (diff={diff})",
                                "one IMEI per physical device",
                                f"Two rows with same AssetId '{aid}' have IMEIs "
                                f"differing by {diff} — may be IMEI1 and IMEI2 of "
                                f"the same device scanned as two listings. "
                                f"Co-rows: {int(r1['co_row'])}, {int(r2['co_row'])}."))
    return flags


# ---------------------------------------------------------------------------
# L5: Duplicates within company data
# ---------------------------------------------------------------------------

def layer5_duplicates(co: pd.DataFrame) -> list[Flag]:
    flags: list[Flag] = []

    dup_imei = (
        co[co["imei"] != ""].groupby("imei").size().loc[lambda s: s > 1]
    )
    for imei, n in dup_imei.items():
        dups = co[co["imei"] == imei]
        for _, r in dups.iterrows():
            flags.append(mk(r, "L5", "duplicate_imei", "CRITICAL",
                            "IMEI Number", imei, "unique per device",
                            f"IMEI '{imei}' appears {n} times "
                            f"(rows {dups['co_row'].tolist()[:5]}); "
                            f"two listings for the same physical device."))

    # AssetId is a product-catalog id — repeats per unit listing are expected.
    # Only flag (AssetId, IMEI) pairs that repeat — that's a true duplicate unit.
    both = co[(co["imei"] != "") & (co["asset_id"] != "")]
    dup_pairs = both.groupby(["asset_id", "imei"]).size().loc[lambda s: s > 1]
    for (aid, imei), n in dup_pairs.items():
        rows = co[(co["asset_id"] == aid) & (co["imei"] == imei)]
        for _, r in rows.iterrows():
            flags.append(mk(r, "L5", "duplicate_asset_id_imei_pair", "CRITICAL",
                            "AssetId + IMEI", f"{aid} / {imei}",
                            "unique (AssetId, IMEI) pair",
                            f"Same AssetId + IMEI pair appears {n} times — "
                            f"this unit was listed more than once."))

    # Same Deal ID with two different IMEIs — one Deal = one device, so two
    # IMEIs under one Deal ID means at least one of them was scanned wrong.
    deal_with_imei = co[(co["appraisal"] != "") & (co["imei"] != "")]
    for deal_id, group in deal_with_imei.groupby("appraisal"):
        unique_imeis = group["imei"].unique()
        if len(unique_imeis) > 1:
            imei_list = list(unique_imeis)[:5]
            for _, r in group.iterrows():
                flags.append(mk(
                    r, "L5", "same_deal_id_multi_imei", "CRITICAL",
                    "Deal ID / IMEI",
                    f"Deal '{deal_id}' / IMEI '{r['imei']}'",
                    "one IMEI per Deal ID",
                    f"Deal ID '{deal_id}' is associated with {len(unique_imeis)} different "
                    f"IMEIs ({imei_list}). One Deal corresponds to one device — at least one "
                    f"of these IMEIs was scanned against the wrong Deal.",
                ))
    return flags


# ---------------------------------------------------------------------------
# L7: Placeholder / test data detection
# ---------------------------------------------------------------------------

PLACEHOLDER_PATTERNS = [
    re.compile(r"^0+$"),                # all zeros
    re.compile(r"^(\d)\1{4,}$"),        # 11111... 22222...
    re.compile(r"^(123456|654321|999999)\d*$"),
    re.compile(r"^TEST", re.I),
    re.compile(r"^DEMO", re.I),
    re.compile(r"^DUMMY", re.I),
]

def looks_like_placeholder(v: str) -> bool:
    if not v:
        return False
    return any(p.match(v) for p in PLACEHOLDER_PATTERNS) or v.startswith("00000")

def layer7_placeholder(co: pd.DataFrame) -> list[Flag]:
    flags: list[Flag] = []
    for _, r in co.iterrows():
        if looks_like_placeholder(r["imei"]):
            flags.append(mk(r, "L7", "placeholder_imei", "HIGH",
                            "IMEI Number", r["imei"], "real device IMEI",
                            "IMEI looks like test/placeholder data — re-scan from device."))
        if looks_like_placeholder(r["barcode"]):
            flags.append(mk(r, "L7", "placeholder_barcode", "MEDIUM",
                            "Barcode", r["barcode"], "real barcode",
                            "Barcode looks like test/placeholder data."))
    return flags


# ---------------------------------------------------------------------------
# L8: Brand validity (Brand column should be a real manufacturer name)
# ---------------------------------------------------------------------------

# Generic / non-brand values seen in the company file that indicate the
# Brand field was filled with a category, OS, or "unknown" placeholder.
INVALID_BRAND_VALUES = {
    "others", "other brands", "other", "windows", "macbooks",
    "smart watch", "smartwatch", "tablet", "laptop", "phone",
    "n/a", "na", "unknown", "tbd", "test",
}

KNOWN_GOOD_BRANDS = (
    set(BRAND_TOKENS.keys()) | {
        "realme", "infinix", "tecno", "zte", "motorola", "blackberry",
        "sony", "nokia", "alcatel", "tcl", "amazon", "fitbit", "bose",
        "jbl", "sonos", "msi", "acer", "razer", "logitech",
    }
)

def layer8_brand_validity(co: pd.DataFrame) -> list[Flag]:
    flags: list[Flag] = []
    for _, r in co.iterrows():
        b = r["brand"]
        if not b:
            flags.append(mk(r, "L8", "brand_missing", "MEDIUM",
                            "Brand", "", "manufacturer name",
                            "Brand column is empty — fill from device or Asset Label."))
            continue
        if b in INVALID_BRAND_VALUES:
            flags.append(mk(r, "L8", "brand_invalid_value", "HIGH",
                            "Brand", b, "manufacturer name (e.g. Apple, Samsung)",
                            f"Brand='{b}' is a category/placeholder, not a manufacturer."))
            continue
        if b not in KNOWN_GOOD_BRANDS:
            # unknown but not obviously bad — low severity heads-up
            flags.append(mk(r, "L8", "brand_unknown", "LOW",
                            "Brand", b, "known manufacturer",
                            f"Brand='{b}' is unfamiliar; verify it's a real manufacturer name."))
    return flags


# ---------------------------------------------------------------------------
# Shared helper: BB catalog index by brand
# ---------------------------------------------------------------------------

def build_brand_idx(catalog: dict) -> dict:
    """brand -> list of (model_str, token_set, data_dict)."""
    idx = defaultdict(list)
    for (b, m), data in catalog.items():
        toks = {t for t in re.split(r"[^a-z0-9]+", m) if len(t) >= 3}
        idx[b].append((m, toks, data))
    return dict(idx)


def best_bb_model(brand: str, label: str, brand_idx: dict):
    """Return (model_str, token_set, data) from BB whose tokens all appear in the label."""
    if not brand or not label or brand not in brand_idx:
        return None
    label_tokens = set(re.split(r"[^a-z0-9]+", label))
    best = None
    for model_str, toks, data in brand_idx[brand]:
        if toks and toks.issubset(label_tokens):
            if best is None or len(toks) > len(best[1]):
                best = (model_str, toks, data)
    return best


# ---------------------------------------------------------------------------
# L9: Identity contradiction — same IMEI, different claimed device
# ---------------------------------------------------------------------------

def layer9_identity_contradiction(co: pd.DataFrame) -> list[Flag]:
    """
    A single IMEI can only belong to ONE physical device. If the same cleaned
    IMEI shows up on multiple rows AND those rows disagree on brand or model
    family, at least one row is wrong — regardless of which row is correct.
    Stronger signal than L5's plain duplicate-IMEI flag.
    """
    flags: list[Flag] = []
    rows_with_imei = co[co["imei"] != ""]
    for imei, group in rows_with_imei.groupby("imei"):
        if len(group) < 2:
            continue
        brands = {b for b in group["brand"].tolist() if b}
        # rough "model family" = first 3 tokens of Asset Label
        fams = set()
        for label in group["asset_label"].tolist():
            toks = [t for t in re.split(r"[^a-z0-9]+", label or "") if len(t) >= 3]
            fams.add(" ".join(toks[:3]))
        if len(brands) > 1 or len(fams) > 1:
            for _, r in group.iterrows():
                flags.append(mk(r, "L9", "imei_identity_contradiction", "CRITICAL",
                                "IMEI Number", imei,
                                "one IMEI = one physical device",
                                f"IMEI '{imei}' appears on {len(group)} rows describing "
                                f"different devices (brands={sorted(brands)}, "
                                f"labels differ). Can't all be correct — one or more is a "
                                f"scan error."))
    return flags


# ---------------------------------------------------------------------------
# L10: TAC cohort anomaly — scanned the wrong phone
# ---------------------------------------------------------------------------

def layer10_tac_cohort(co: pd.DataFrame, brand_idx: dict) -> list[Flag]:
    """
    TAC = first 8 digits of an IMEI, encoding make+model. Within a cohort of
    rows claiming the same (brand, BB-matched model), the TAC should be
    consistent. If one row's TAC is a singleton while others in the cohort
    share a TAC (≥3 occurrences), the operator likely scanned the IMEI off
    the wrong phone on the shelf.
    """
    flags: list[Flag] = []
    valid = co[(co["imei_shape"] == "imei15")].copy()
    if len(valid) < 5:
        return flags
    valid["tac"] = valid["imei"].str[:8]

    def cohort_key(row):
        m = best_bb_model(row["brand"], row["asset_label"], brand_idx)
        return (row["brand"], m[0] if m else "")

    valid["cohort"] = valid.apply(cohort_key, axis=1)

    for cohort, group in valid.groupby("cohort"):
        if cohort[0] == "" or cohort[1] == "" or len(group) < 5:
            continue
        tac_counts = Counter(group["tac"].tolist())
        # need at least one "dominant" TAC for this cohort to be meaningful
        dominant = [t for t, c in tac_counts.items() if c >= 3]
        if not dominant:
            continue
        for _, r in group.iterrows():
            if tac_counts[r["tac"]] == 1:
                flags.append(mk(r, "L10", "tac_cohort_anomaly", "HIGH",
                                "IMEI Number", r["imei"],
                                f"TAC matching cohort {cohort[0]} {cohort[1]}",
                                f"Every other {cohort[0]} {cohort[1]} row in this batch "
                                f"has an IMEI starting with one of {dominant}; this row's "
                                f"IMEI starts with {r['tac']}. Likely the operator scanned "
                                f"a different phone. Re-scan to confirm."))
    return flags


# ---------------------------------------------------------------------------
# L11: Model-number mismatch against BB catalog
# ---------------------------------------------------------------------------

APPLE_MODEL_NUM = re.compile(r"\bA\d{4}\b", re.I)
SAMSUNG_MODEL_NUM = re.compile(r"\bSM-[A-Z0-9]{4,8}\b", re.I)

def _extract_model_numbers(label: str) -> list[str]:
    out = []
    out += [m.group(0).upper() for m in APPLE_MODEL_NUM.finditer(label)]
    out += [m.group(0).upper() for m in SAMSUNG_MODEL_NUM.finditer(label)]
    return out

def layer11_model_number(co: pd.DataFrame, brand_idx: dict) -> list[Flag]:
    flags: list[Flag] = []
    for _, r in co.iterrows():
        label = r["asset_label"] or ""
        found = _extract_model_numbers(label)
        if not found:
            continue
        bb_match = best_bb_model(r["brand"], label, brand_idx)
        if not bb_match:
            continue
        model_str, _, data = bb_match
        bb_numbers = {n.upper() for n in data.get("model_numbers", set()) if n}
        if not bb_numbers:
            continue
        for mn in found:
            if mn not in bb_numbers and not any(mn in bbn or bbn in mn for bbn in bb_numbers):
                flags.append(mk(r, "L11", "model_number_mismatch", "HIGH",
                                "Asset Label (model number)", mn,
                                f"one of {sorted(bb_numbers)}",
                                f"Label claims model '{model_str}' but mentions model "
                                f"number '{mn}', which Blackbelt has never seen for that "
                                f"model. Either the model name or the model number is wrong."))
                break  # one flag per row is enough
    return flags


# ---------------------------------------------------------------------------
# L12: Color not in BB catalog for this model
# ---------------------------------------------------------------------------

KNOWN_COLORS = {
    "black", "white", "silver", "gold", "rose", "blue", "red", "green",
    "purple", "pink", "yellow", "orange", "grey", "gray", "graphite",
    "titanium", "midnight", "starlight", "sierra", "pacific", "alpine",
    "natural", "desert", "space",
}

def _color_tokens_in_label(label: str) -> set[str]:
    toks = set(re.split(r"[^a-z]+", label or ""))
    return toks & KNOWN_COLORS

def layer12_color_catalog(co: pd.DataFrame, brand_idx: dict) -> list[Flag]:
    flags: list[Flag] = []
    for _, r in co.iterrows():
        label = r["asset_label"] or ""
        co_colors = _color_tokens_in_label(label)
        if not co_colors:
            continue
        bb_match = best_bb_model(r["brand"], label, brand_idx)
        if not bb_match:
            continue
        model_str, _, data = bb_match
        bb_colors_text = " ".join(data.get("colors", set()))
        bb_color_tokens = set(re.split(r"[^a-z]+", bb_colors_text)) & KNOWN_COLORS
        if not bb_color_tokens:
            continue
        unseen = co_colors - bb_color_tokens
        if unseen:
            flags.append(mk(r, "L12", "color_not_in_bb_catalog", "MEDIUM",
                            "Asset Label (color)", ", ".join(sorted(unseen)),
                            f"one of {sorted(bb_color_tokens)}",
                            f"Label mentions color(s) {sorted(unseen)} for {r['brand']} "
                            f"{model_str}, but Blackbelt has only ever recorded this model "
                            f"in {sorted(bb_color_tokens)}. May be a rare variant — verify."))
    return flags


# ---------------------------------------------------------------------------
# L13: Two different storage values in one label
# ---------------------------------------------------------------------------

STORAGE_ALL = re.compile(r"(\d+)\s*(TB|GB)\b", re.I)

def layer13_two_storages(co: pd.DataFrame) -> list[Flag]:
    flags: list[Flag] = []
    for _, r in co.iterrows():
        label = r["asset_label"] or ""
        matches = STORAGE_ALL.findall(label)
        if len(matches) < 2:
            continue
        # normalise to GB
        sizes = set()
        for n, unit in matches:
            sizes.add(int(n) * 1024 if unit.upper() == "TB" else int(n))
        if len(sizes) >= 2:
            flags.append(mk(r, "L13", "two_storages_in_label", "MEDIUM",
                            "Asset Label", label,
                            "single storage value",
                            f"Label mentions two different storage sizes "
                            f"({sorted(sizes)}GB) — likely two listings merged, or "
                            f"template partially overwritten. Keep only the correct one."))
    return flags


# ---------------------------------------------------------------------------
# L14: Grade contradicts damage keywords in the label
# ---------------------------------------------------------------------------

DAMAGE_KEYWORDS = [
    "faulty", "dead", "no power", "not working", "cracked", "broken",
    "damaged", "water damage", "liquid damage", "shattered", "defective",
    "screen broken", "back glass broken", "wont turn on", "won't turn on",
]
GOOD_GRADES = {"a", "a+", "a grade", "excellent", "good", "new", "mint", "pristine", "fresh"}

def layer14_grade_damage(co: pd.DataFrame) -> list[Flag]:
    flags: list[Flag] = []
    for _, r in co.iterrows():
        label = r["asset_label"] or ""
        grade = (r["grade"] or "").strip()
        if grade not in GOOD_GRADES:
            continue
        found = [kw for kw in DAMAGE_KEYWORDS if kw in label]
        if found:
            flags.append(mk(r, "L14", "grade_contradicts_damage", "MEDIUM",
                            "Grade", grade,
                            "damaged-tier grade (B/C/Faulty) given the label",
                            f"Label mentions damage ({found}) but Grade='{grade}'. "
                            f"Either the label is from a different unit, or the grade was "
                            f"entered wrong."))
    return flags


# ---------------------------------------------------------------------------
# L15: QR column contradicts IMEI column
# ---------------------------------------------------------------------------

def layer15_qr_vs_imei(co: pd.DataFrame) -> list[Flag]:
    flags: list[Flag] = []
    for _, r in co.iterrows():
        qr_cleaned = clean_id(r["qr"])
        if imei_shape(qr_cleaned) != "imei15":
            continue
        if not r["imei"] or qr_cleaned == r["imei"]:
            continue
        # Only flag when BOTH look like real IMEIs — otherwise QR is just something else.
        if luhn_valid(qr_cleaned) and luhn_valid(r["imei"]):
            flags.append(mk(r, "L15", "qr_code_contradicts_imei", "MEDIUM",
                            "QR Code / IMEI Number",
                            f"QR='{qr_cleaned}', IMEI='{r['imei']}'",
                            "both should match (or QR should hold a non-IMEI value)",
                            "QR column holds a valid IMEI that differs from the IMEI "
                            "column. One of them was scanned from the wrong phone."))
    return flags


# ---------------------------------------------------------------------------
# L16: (brand, model) not in BB catalog at all — advisory
# ---------------------------------------------------------------------------

def layer16_catalog_gap(co: pd.DataFrame, brand_idx: dict) -> list[Flag]:
    """
    Advisory only. If the CO row's brand is covered by BB but no BB model
    matches the label, Blackbelt has no reference data to validate against.
    This is not an error per se — it's a request to extend BB's catalog.
    """
    flags: list[Flag] = []
    covered = set(brand_idx)
    for _, r in co.iterrows():
        b, label = r["brand"], r["asset_label"]
        if not b or not label or b not in covered:
            continue
        if best_bb_model(b, label, brand_idx) is None:
            flags.append(mk(r, "L16", "brand_model_not_in_bb_catalog", "LOW",
                            "Asset Label", label,
                            "a Blackbelt-catalog model",
                            f"Blackbelt has {b} in its catalog but no model matching "
                            f"this label. May be a legitimate new model — consider adding "
                            f"it to Blackbelt so future scans can be validated."))
    return flags


# ---------------------------------------------------------------------------
# L18: BB ↔ Master field-level reconciliation
# When the Master row's IMEI matches a BB row, compare field-by-field.
# BB is treated as the truth — disagreements are flagged HIGH (Confirmed
# Errors) because BB's data comes from a physical device test, not from
# manual entry.
# ---------------------------------------------------------------------------

def _grade_normalize(g: str) -> str:
    """Normalize grade strings for comparison.
    BB emits A / A+ / B / C / D1 / D2 / F. Stack emits 'Grade A',
    'Grade A-Plus', 'Grade D2.' etc. We strip the 'Grade ' prefix and trailing
    punctuation, then map A-Plus aliases onto 'a+' so A and A+ stay distinct.
    """
    g = norm_text(g)
    g = re.sub(r"^grade\s+", "", g)
    g = re.sub(r"[.\s]+$", "", g).strip()  # strip trailing dot/space, KEEP '+'
    aliases = {"a": "a", "a+": "a+",
               "a-plus": "a+", "a plus": "a+", "aplus": "a+",
               "as new": "a", "as-new": "a", "new": "a",
               "b": "b", "c": "c", "d": "d", "d1": "d1", "d2": "d2",
               "e": "e", "f": "f"}
    return aliases.get(g, g)


def layer18_bb_reconciliation(co: pd.DataFrame, bb_records: dict) -> list[Flag]:
    """
    Per-row field reconciliation against Blackbelt's recorded test results.
    Skipped on rows whose IMEI is not in BB (no record to compare against).
    """
    flags: list[Flag] = []
    if not bb_records:
        return flags
    for _, r in co.iterrows():
        rec = bb_records.get(str(r["imei"]))
        if rec is None:
            continue

        # Brand
        bb_brand, m_brand = rec["brand"], r["brand"]
        if bb_brand and m_brand and bb_brand != m_brand:
            flags.append(mk(r, "L18", "bb_brand_mismatch", "CRITICAL", "Brand",
                            f"Master='{m_brand}', Blackbelt='{bb_brand}'",
                            f"'{bb_brand}' (per Blackbelt test)",
                            f"Blackbelt physically tested this IMEI as a {bb_brand} "
                            f"device, but Master records it as {m_brand}. Update "
                            f"Master.Brand to '{bb_brand}'."))

        # Model — check that BB's model tokens are a subset of Master's label
        bb_model, m_label = rec["model"], r["asset_label"]
        if bb_model and m_label:
            bb_tokens = {t for t in re.split(r"[^a-z0-9]+", bb_model) if len(t) >= 3}
            m_tokens  = set(re.split(r"[^a-z0-9]+", m_label))
            if bb_tokens and not bb_tokens.issubset(m_tokens):
                flags.append(mk(r, "L18", "bb_model_mismatch", "HIGH", "Model",
                                f"Master='{m_label}', Blackbelt='{bb_model}'",
                                f"'{bb_model}' (per Blackbelt test)",
                                f"Blackbelt tested this IMEI as '{bb_model}', but "
                                f"Master's label doesn't mention it. Update Master."))

        # Storage
        bb_st, m_st = rec["storage_gb"], r["storage_gb"]
        if bb_st and m_st and int(bb_st) != int(m_st):
            flags.append(mk(r, "L18", "bb_storage_mismatch", "HIGH", "Storage",
                            f"Master={int(m_st)}GB, Blackbelt={int(bb_st)}GB",
                            f"{int(bb_st)}GB (per Blackbelt test)",
                            f"Blackbelt's hardware read confirms {int(bb_st)}GB, but "
                            f"Master records {int(m_st)}GB. Affects price; correct Master."))

        # Grade
        bb_g, m_g = _grade_normalize(rec["grade"]), _grade_normalize(r["grade"])
        if bb_g and m_g and bb_g != m_g:
            flags.append(mk(r, "L18", "bb_grade_mismatch", "HIGH", "Grade",
                            f"Master='{r['grade']}', Blackbelt='{rec['grade']}'",
                            f"'{rec['grade']}' (per Blackbelt test machine)",
                            f"Blackbelt's automated grading disagrees with Master. "
                            f"BB's grade is the test-machine output; consider it "
                            f"authoritative. Update Master."))

        # Color (only if BB has it AND Master's label clearly doesn't include it)
        bb_c = rec["color"].lower() if rec["color"] else ""
        if bb_c and m_label:
            color_tokens = set(re.split(r"[^a-z]+", bb_c))
            label_tokens = set(re.split(r"[^a-z]+", m_label))
            if color_tokens and not (color_tokens & label_tokens):
                flags.append(mk(r, "L18", "bb_color_mismatch", "MEDIUM", "Color",
                                f"Master='{m_label}', Blackbelt color='{rec['color']}'",
                                f"label mentioning '{rec['color']}'",
                                f"Blackbelt recorded color '{rec['color']}'; Master's "
                                f"asset label doesn't mention it. Verify."))

        # Model number (Apple A-codes, Samsung SM-codes — when present in BB)
        bb_mn = (rec["model_number"] or "").upper().replace(" ", "")
        if bb_mn and len(bb_mn) >= 4 and m_label:
            label_codes = re.findall(r"(?:A\d{4,5}|SM-[A-Z0-9]+)", m_label.upper())
            if label_codes and bb_mn not in label_codes:
                flags.append(mk(r, "L18", "bb_model_number_mismatch", "MEDIUM",
                                "Model number",
                                f"Master mentions {label_codes}, Blackbelt has '{bb_mn}'",
                                f"label including '{bb_mn}'",
                                f"Blackbelt recorded model number '{bb_mn}', but "
                                f"Master's label has {label_codes}. Verify regional variant."))
    return flags


# ---------------------------------------------------------------------------
# L19: Master ↔ Stack Bulk cross-validation
# Stack Bulk is the sell-side record of the same physical inventory. When
# Master and Stack disagree on grade / VAT / country / IMEI presence, at
# least one side needs correcting.
# ---------------------------------------------------------------------------

def layer19_master_stack_recon(co: pd.DataFrame,
                               stack_by_imei: dict,
                               stack_by_deal: dict) -> list[Flag]:
    flags: list[Flag] = []
    if not stack_by_imei and not stack_by_deal:
        return flags
    for _, r in co.iterrows():
        deal = str(r.get("appraisal", "") or "").strip()
        s_by_imei = stack_by_imei.get(str(r["imei"])) if r["imei"] else None
        s_by_deal = stack_by_deal.get(deal) if deal else None

        # Grade mismatch (when both sides have a grade and rows can be joined by IMEI)
        if s_by_imei:
            mg, sg = _grade_normalize(r["grade"]), _grade_normalize(s_by_imei.get("grade", ""))
            if mg and sg and mg != sg:
                flags.append(mk(r, "L19", "master_stack_grade_mismatch", "MEDIUM",
                                "Grade",
                                f"Master='{r['grade']}', Stack='{s_by_imei['grade']}'",
                                "Master and Stack to agree on grade",
                                f"Master and Stack-Bulk record different grades for the "
                                f"same IMEI. Affects pricing — reconcile with the grader."))

            # VAT mismatch
            mv = norm_text(r.get("vat_type", ""))
            sv = norm_text(s_by_imei.get("vat_type", ""))
            if mv and sv and mv != sv:
                flags.append(mk(r, "L19", "master_stack_vat_mismatch", "MEDIUM",
                                "VAT Type",
                                f"Master='{r['vat_type']}', Stack='{s_by_imei['vat_type']}'",
                                "Master and Stack to agree on VAT type",
                                "Master and Stack disagree on VAT classification — "
                                "affects pricing & tax. Reconcile."))

            # Country mismatch
            mc = norm_text(r.get("country", ""))
            sc = norm_text(s_by_imei.get("country", ""))
            if mc and sc and mc != sc:
                flags.append(mk(r, "L19", "master_stack_country_mismatch", "LOW",
                                "Country",
                                f"Master='{r['country']}', Stack='{s_by_imei['country']}'",
                                "Master and Stack to agree on storage country",
                                "Master and Stack disagree on storage country — "
                                "either inventory moved or one record is stale."))

        # IMEI not in Stack at all (but Deal is) — Master IMEI may be wrong
        if r["imei"] and not s_by_imei and s_by_deal:
            stack_imei = str(s_by_deal.get("imei", "") or "").strip()
            if stack_imei and stack_imei != str(r["imei"]):
                flags.append(mk(r, "L19", "master_imei_disagrees_with_stack", "HIGH",
                                "IMEI Number",
                                f"Master='{r['imei']}', Stack has '{stack_imei}' for the same Deal",
                                f"'{stack_imei}' (per Stack Bulk for this Deal)",
                                f"Master's IMEI doesn't match Stack's IMEI for the "
                                f"same Deal ID '{deal}'. Stack records '{stack_imei}'; "
                                f"verify which is correct and update Master."))

        # IMEI in Master but Stack has no record at all (no IMEI, no Deal)
        if r["imei"] and not s_by_imei and not s_by_deal:
            flags.append(mk(r, "L19", "master_not_in_stack", "LOW",
                            "IMEI / Deal ID",
                            f"IMEI='{r['imei']}', Deal='{deal}'",
                            "device should appear in Stack Bulk if queued for sale",
                            "Device exists in Master but isn't in the Stack Bulk file. "
                            "Either it hasn't been queued for sale yet, or one of the "
                            "two records is wrong."))
    return flags


# ---------------------------------------------------------------------------
# L20: Stale inventory — Deal date is older than 12 months
# ---------------------------------------------------------------------------

def layer20_stale_inventory(co: pd.DataFrame, today_iso: str = None) -> list[Flag]:
    from datetime import date
    today = date.fromisoformat(today_iso) if today_iso else date.today()
    flags: list[Flag] = []
    for _, r in co.iterrows():
        d = r.get("deal_date")
        if not d:
            continue
        try:
            deal_d = date.fromisoformat(d)
        except (ValueError, TypeError):
            continue
        days = (today - deal_d).days
        if days > 365:
            months = days // 30
            flags.append(mk(r, "L20", "stale_inventory", "LOW",
                            "Deal date",
                            f"Deal opened {deal_d.isoformat()} ({months} months ago)",
                            "active inventory normally turns over within 12 months",
                            f"This unit has been in inventory for ~{months} months. "
                            f"Review whether to liquidate, return, or keep listing."))
    return flags


# ---------------------------------------------------------------------------
# L21: BB test-result failures
# When BB's recorded tests have FAIL outcomes (Battery, Bluetooth, Screen,
# etc.), surface them so the inventory team knows the device shouldn't be
# sold as-is.
# ---------------------------------------------------------------------------

def layer21_bb_test_failures(co: pd.DataFrame, bb_records: dict) -> list[Flag]:
    flags: list[Flag] = []
    if not bb_records:
        return flags
    for _, r in co.iterrows():
        rec = bb_records.get(str(r["imei"]))
        if not rec:
            continue
        failed = [name for name, val in rec.get("tests", {}).items()
                  if val == "FAIL"]
        if not failed:
            continue
        flags.append(mk(r, "L21", "bb_test_failed", "HIGH",
                        "Blackbelt test results",
                        f"Failed: {', '.join(failed[:5])}"
                        + (f" (+{len(failed)-5} more)" if len(failed) > 5 else ""),
                        "all Blackbelt tests passed",
                        f"Blackbelt's automated test machine recorded {len(failed)} "
                        f"failed test(s) for this device. Do not list as fully "
                        f"functional until repaired/regraded."))
    return flags


# ---------------------------------------------------------------------------
# L22: BB refurbished-part detection
# When BB's RP (Replaced Part) columns flag a non-genuine component, surface
# it. This is a material disclosure item for resale.
# ---------------------------------------------------------------------------

def layer22_bb_refurbished_parts(co: pd.DataFrame, bb_records: dict) -> list[Flag]:
    flags: list[Flag] = []
    if not bb_records:
        return flags
    for _, r in co.iterrows():
        rec = bb_records.get(str(r["imei"]))
        if not rec:
            continue
        parts = rec.get("refurbished", [])
        if not parts:
            continue
        flags.append(mk(r, "L22", "bb_refurbished_parts", "MEDIUM",
                        "Blackbelt parts check",
                        f"Non-genuine: {', '.join(parts[:5])}"
                        + (f" (+{len(parts)-5} more)" if len(parts) > 5 else ""),
                        "all components genuine",
                        f"Blackbelt detected {len(parts)} non-genuine/replaced "
                        f"component(s). Material disclosure for resale; verify "
                        f"price tier and customer-facing description."))
    return flags


# ---------------------------------------------------------------------------
# L17: Blackbelt coverage gap — IMEI is well-formed but not in the BB file
# ---------------------------------------------------------------------------

def layer17_blackbelt_coverage(co: pd.DataFrame, bb_by_imei: dict) -> list[Flag]:
    """
    Advisory only. Flags rows whose IMEI is a Luhn-valid 15-digit IMEI but
    does not appear in the Blackbelt file's IMEI/IMEI2 columns. Surfaces
    the coverage gap so the manager can see which inventory has not been
    tested by Blackbelt and request additional testing.

    Skipped if no Blackbelt data was provided (empty bb_by_imei).
    Skipped on rows whose IMEI is malformed (those already get a higher-
    severity L1 flag, so the missing BB record is not the actionable
    issue).
    """
    flags: list[Flag] = []
    if not bb_by_imei:
        return flags
    bb_keys = set(bb_by_imei.keys())
    for _, r in co.iterrows():
        if r["imei_shape"] != "imei15":
            continue
        if not luhn_valid(r["imei"]):
            continue
        if r["imei"] in bb_keys:
            continue
        flags.append(mk(
            r, "L17", "not_in_blackbelt", "LOW", "IMEI Number",
            r["imei"], "Blackbelt has tested this device",
            "Valid IMEI is not present in the Blackbelt reference file. "
            "Blackbelt has not tested this device — consider including this "
            "model/batch in the next round of Blackbelt testing.",
        ))
    return flags


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

# CSV filenames (kept for internal use)
UI_FILE_HIGH      = "high_confidence_matches.csv"
UI_FILE_MEDIUM    = "medium_confidence_matches.csv"
UI_FILE_LOW       = "low_confidence_matches.csv"
UI_FILE_UNMATCHED = "unmatched.csv"
UI_FILE_SUMMARY   = "summary.json"

# Excel filenames — these are what users actually download
XLSX_FILE_HIGH      = "verified_matches.xlsx"
XLSX_FILE_MEDIUM    = "likely_matches.xlsx"
XLSX_FILE_LOW       = "uncertain_matches.xlsx"
XLSX_FILE_UNMATCHED = "clean_rows.xlsx"
XLSX_FILE_GRADES    = "grade_mismatches.xlsx"


# ---------------------------------------------------------------------------
# Human-readable mappings for the Excel report
# ---------------------------------------------------------------------------

# Per-issue metadata: (plain-English problem name, concise fix, plain expected)
ISSUE_INFO: dict[str, dict[str, str]] = {
    # L1 FORMAT
    "imei_missing":      {"problem": "IMEI is missing",
                          "fix": "Scan the phone (dial *#06#) and fill the IMEI column.",
                          "expected": "A 15-digit IMEI from the device."},
    "imei_luhn_fail":    {"problem": "Mobile-phone IMEI fails the Luhn checksum (industry-standard digit verification)",
                          "fix": "Re-scan the IMEI from the device (dial *#06#). The current value cannot be a real IMEI — see the legend sheet for how the Luhn check works. (Rule applies to mobile phones only.)",
                          "expected": "A valid 15-digit IMEI whose check digit (last digit) satisfies the Luhn formula. Legend sheet has a worked example."},
    "imei_wrong_length": {"problem": "Mobile-phone IMEI is not 15 digits",
                          "fix": "Re-scan the IMEI from the device. (Rule applies to mobile phones only — for laptops, tablets, smartwatches, earphones the column may legitimately hold a manufacturer serial number of any length; see the legend sheet.)",
                          "expected": "A 15-digit numeric IMEI for mobile phones."},
    "looks_like_imeisv": {"problem": "16-digit value looks like an IMEISV instead of a 15-digit IMEI",
                          "fix": "IMEISV is the IMEI plus a 2-digit Software Version code. The first 14 digits of this value form a Luhn-valid IMEI body — drop the trailing 2 digits and replace with the standard 15-digit IMEI, or leave as-is if your downstream system tracks IMEISV intentionally.",
                          "expected": "A 15-digit IMEI. The first 14 digits of this value are valid; only the trailing 2 digits make it 16 long."},
    # L2 SCAN-SLOT
    "imei_in_barcode_slot": {"problem": "IMEI was entered into the Barcode column",
                             "fix": "Move this value to IMEI; re-scan the internal barcode.",
                             "expected": "An internal barcode, not an IMEI."},
    "imei_equals_barcode":  {"problem": "IMEI and Barcode hold the same value",
                             "fix": "Fill the Barcode column with the real barcode.",
                             "expected": "Different values in IMEI vs Barcode."},
    # L3 INTRA-ROW
    "brand_label_contradiction": {"problem": "Brand doesn't match the product in the Asset Label",
                                  "fix": "Correct the Brand field to match the label.",
                                  "expected": "Brand consistent with the Asset Label."},
    "brand_token_absent":        {"problem": "Asset Label doesn't mention the claimed brand",
                                  "fix": "Verify the Brand and Asset Label agree.",
                                  "expected": "Label that mentions the brand's product family."},
    "category_model_mismatch":   {"problem": "Category doesn't match the product name",
                                  "fix": "Set Category to match the label (e.g. iPad → Tablet).",
                                  "expected": "Category consistent with the product."},
    "storage_missing":           {"problem": "Storage size is missing from the Asset Label",
                                  "fix": "Add storage to the Asset Label (e.g. '128GB').",
                                  "expected": "Asset Label including a storage size."},
    # L4
    "storage_unseen_in_bb": {"problem": "Storage size not recorded by Blackbelt for this model",
                             "fix": "Verify the label — may be a rare variant or a typo.",
                             "expected": "Storage size Blackbelt has seen for this model."},
    # L5
    "duplicate_imei":                {"problem": "This IMEI appears on multiple rows",
                                      "fix": "Keep one row; delete the duplicates.",
                                      "expected": "Each IMEI appears on exactly one row."},
    "duplicate_asset_id_imei_pair":  {"problem": "Same Asset ID + IMEI listed twice",
                                      "fix": "Delete the duplicate listing.",
                                      "expected": "One listing per (Asset ID, IMEI) pair."},
    "same_deal_id_multi_imei":       {"problem": "One Deal ID is associated with two or more different IMEIs",
                                      "fix": "A Deal ID identifies a single device transaction, so it must map to exactly one IMEI. Compare the listed IMEIs against the device(s) physically held under this Deal and keep the correct IMEI; the others were scanned against the wrong Deal.",
                                      "expected": "Exactly one IMEI per Deal ID."},
    # L6
    "possible_imei1_imei2_pair": {"problem": "Same phone listed twice using IMEI1 and IMEI2",
                                  "fix": "Keep one row per physical device.",
                                  "expected": "One row per phone, not one row per IMEI."},
    # L7
    "placeholder_imei":    {"problem": "Fake / test IMEI in production data",
                            "fix": "Re-scan the real IMEI from the device.",
                            "expected": "A real 15-digit IMEI from the device."},
    "placeholder_barcode": {"problem": "Fake / test barcode in production data",
                            "fix": "Replace with the real barcode.",
                            "expected": "A real barcode from the device."},
    # L8
    "brand_missing":       {"problem": "Brand field is empty",
                            "fix": "Fill Brand with the manufacturer name (e.g. Apple, Samsung).",
                            "expected": "A real manufacturer name."},
    "brand_invalid_value": {"problem": "Brand column holds a category, not a manufacturer",
                            "fix": "Replace with the actual manufacturer name.",
                            "expected": "A manufacturer like Apple, Samsung, etc."},
    "brand_unknown":       {"problem": "Brand isn't a recognised manufacturer",
                            "fix": "Verify the Brand field is a real manufacturer.",
                            "expected": "A known manufacturer name."},
    # L9
    "imei_identity_contradiction": {"problem": "Same IMEI claimed on two different devices",
                                    "fix": "Investigate both rows — only one can be correct.",
                                    "expected": "Each IMEI belongs to one physical device."},
    # L10
    "tac_cohort_anomaly": {"problem": "IMEI looks like it was scanned off a different phone",
                           "fix": "Re-scan the IMEI off the device on hand.",
                           "expected": "An IMEI that starts with this model's typical prefix."},
    # L11
    "model_number_mismatch": {"problem": "Model code in label doesn't belong to the named model",
                              "fix": "Fix either the model name or the model code.",
                              "expected": "A model code matching the named model."},
    # L12
    "color_not_in_bb_catalog": {"problem": "Colour not recorded by Blackbelt for this model",
                                "fix": "Verify the colour in the label.",
                                "expected": "A colour Blackbelt has seen for this model."},
    # L13
    "two_storages_in_label": {"problem": "Asset Label has two different storage sizes",
                              "fix": "Keep only the correct storage size.",
                              "expected": "A single storage size in the label."},
    # L14
    "grade_contradicts_damage": {"problem": "Grade says excellent/new but label mentions damage",
                                 "fix": "Fix the Grade or the Asset Label to agree.",
                                 "expected": "Grade that matches the device condition."},
    # L15
    "qr_code_contradicts_imei": {"problem": "QR code is a different IMEI than the IMEI column",
                                 "fix": "Re-scan — one value is from the wrong phone.",
                                 "expected": "QR and IMEI columns hold the same IMEI."},
    # L16
    "brand_model_not_in_bb_catalog": {"problem": "Model not found in Blackbelt's catalog",
                                      "fix": "Ask Blackbelt to add this model to their catalog.",
                                      "expected": "A model with Blackbelt reference data."},
    # L17
    "not_in_blackbelt":              {"problem": "Device has not been tested by Blackbelt",
                                      "fix": "The IMEI is well-formed but no Blackbelt test record exists for this device. Include the unit (or its model/batch) in the next round of Blackbelt testing.",
                                      "expected": "An IMEI that appears in the Blackbelt reference file."},
    # L18 — BB ↔ Master field reconciliation (BB is treated as truth)
    "bb_brand_mismatch":             {"problem": "Brand in Master disagrees with Blackbelt's recorded brand",
                                      "fix": "Blackbelt's hardware test identified a different manufacturer than what Master records. Update Master.Brand to match Blackbelt — the test machine reads from the device firmware.",
                                      "expected": "Master and Blackbelt to agree on Brand."},
    "bb_model_mismatch":             {"problem": "Model in Master doesn't match Blackbelt's recorded model",
                                      "fix": "Blackbelt's recorded model name is not present in Master's asset label. Update Master to match Blackbelt's model.",
                                      "expected": "Master's label to mention the model Blackbelt recorded."},
    "bb_storage_mismatch":           {"problem": "Storage size in Master disagrees with Blackbelt's recorded storage",
                                      "fix": "Blackbelt's hardware read disagrees with Master's storage size. Storage drives price by 10–30%; correct Master to match Blackbelt's value.",
                                      "expected": "Master and Blackbelt to agree on storage size."},
    "bb_grade_mismatch":             {"problem": "Grade in Master disagrees with Blackbelt's automated grade",
                                      "fix": "Blackbelt's grade comes from the test machine and is more rigorous than manual grading. Treat BB's grade as authoritative; update Master.",
                                      "expected": "Master and Blackbelt to agree on Grade."},
    "bb_color_mismatch":             {"problem": "Color in Master's label disagrees with Blackbelt's recorded color",
                                      "fix": "Blackbelt recorded a different colour. Verify the asset label and update Master.",
                                      "expected": "Master's label to mention Blackbelt's recorded colour."},
    "bb_model_number_mismatch":      {"problem": "Model number code disagrees between Master's label and Blackbelt",
                                      "fix": "Master's label has a different manufacturer code than Blackbelt recorded. May indicate a different regional variant — verify.",
                                      "expected": "Master's label model-number code to match Blackbelt."},
    # L19 — Master ↔ Stack reconciliation
    "master_stack_grade_mismatch":   {"problem": "Master and Stack-Bulk record different grades for the same IMEI",
                                      "fix": "Two separate graders disagree on the device's grade. Affects price; reconcile with the team and use the more rigorous source.",
                                      "expected": "Master and Stack to agree on Grade."},
    "master_stack_vat_mismatch":     {"problem": "Master and Stack record different VAT classifications",
                                      "fix": "VAT type drives tax handling. Reconcile and update the wrong side.",
                                      "expected": "Master and Stack to agree on VAT Type."},
    "master_stack_country_mismatch": {"problem": "Master and Stack record different storage countries",
                                      "fix": "Either inventory moved between locations or one record is stale. Update the lagging record.",
                                      "expected": "Master and Stack to agree on storage country."},
    "master_imei_disagrees_with_stack": {"problem": "Master's IMEI doesn't match Stack's IMEI for the same Deal",
                                      "fix": "The same Deal ID has one IMEI in Master and a different IMEI in Stack. Stack's value is shown — verify which is correct and update Master.",
                                      "expected": "Master and Stack to record the same IMEI for one Deal."},
    "master_not_in_stack":           {"problem": "Device exists in Master but no record in Stack Bulk",
                                      "fix": "Either the device hasn't been queued for sale yet, or one of the records is wrong. Investigate.",
                                      "expected": "Master devices that should be on sale to also exist in Stack."},
    # L20 — stale inventory
    "stale_inventory":               {"problem": "Device has been in inventory for more than 12 months",
                                      "fix": "Long-held inventory ties up capital. Decide whether to liquidate, return, or continue listing.",
                                      "expected": "Active inventory typically turns over within 12 months."},
    # L21 — BB test failures
    "bb_test_failed":                {"problem": "Blackbelt's test machine recorded one or more failed hardware tests",
                                      "fix": "The device is not fully functional per Blackbelt's automated test. Do not list as fully working until repaired and re-tested.",
                                      "expected": "All Blackbelt hardware tests to pass."},
    # L22 — refurbished parts
    "bb_refurbished_parts":          {"problem": "Blackbelt detected one or more non-genuine / replaced components",
                                      "fix": "The device has been previously repaired with non-OEM parts. Material disclosure for resale; adjust price tier and update customer-facing description.",
                                      "expected": "All components to read as genuine in Blackbelt's parts check."},
    "serial_in_imei_slot":           {"problem": "Serial number entered into the IMEI column",
                                      "fix": "An alphanumeric serial number was scanned into the IMEI slot. Re-scan the IMEI from the device (dial *#06# on phones; on tablets check the device label).",
                                      "expected": "A 15-digit numeric IMEI in the IMEI column."},
}

# Issues that don't indicate a per-row data error — coverage gaps, catalog
# gaps, format-variant advisories, and operational signals. Rows whose ONLY
# flags are in this set are still considered "clean" (no real problem) so
# they populate the All-good download instead of being swept into LOW.
# The recommendations sheet still surfaces these as informational items.
ADVISORY_ISSUES: set[str] = {
    "not_in_blackbelt",              # L17 — BB hasn't tested this device yet
    "brand_model_not_in_bb_catalog", # L16 — BB has no reference for this model
    "looks_like_imeisv",             # L1  — legitimate IMEISV (16-digit) format
    "stale_inventory",               # L20 — operational, not a data error
}

# Per-layer metadata for the 'How to Read This Report' guide sheet
LAYER_INFO: list[tuple[str, str, str]] = [
    ("L1",  "Format check",           "Is the IMEI the right shape (15 digits, passes digit-check)?"),
    ("L2",  "Wrong column",           "Is a value in the wrong column (serial in IMEI, IMEI in barcode)?"),
    ("L3",  "Internal consistency",   "Do Brand, Asset Label, and Category agree with each other?"),
    ("L4",  "Storage catalog check", "Does the storage size exist for this model in Blackbelt?"),
    ("L5",  "Duplicate rows",         "Is the same IMEI or same unit listed more than once?"),
    ("L6",  "IMEI1/IMEI2 twin",       "Are dual-SIM IMEIs of one phone listed as two rows?"),
    ("L7",  "Placeholder data",       "Is the IMEI a test / fake / all-zeros value?"),
    ("L8",  "Brand column sanity",    "Is the Brand column filled with a real manufacturer?"),
    ("L9",  "Identity contradiction", "Same IMEI claimed as two different devices?"),
    ("L10", "Scanned the wrong phone","Does the IMEI prefix match others of this model?"),
    ("L11", "Model code match",       "Does the Apple/Samsung code in the label match the model?"),
    ("L12", "Colour catalog",         "Has Blackbelt recorded this model in this colour?"),
    ("L13", "Storage confusion",      "Does the label mention two different storage sizes?"),
    ("L14", "Grade vs damage",        "Is a damaged phone graded as excellent?"),
    ("L15", "QR vs IMEI",             "Do QR code and IMEI column disagree?"),
    ("L16", "Unknown model",          "Is this model missing from Blackbelt's catalog?"),
]

PRIORITY_DISPLAY = {
    "CRITICAL": "1 - Most Important",
    "HIGH":     "2 - Important",
    "MEDIUM":   "3 - Review",
    "LOW":      "4 - Advisory",
}

PRIORITY_EXPLAIN = [
    ("1 - Most Important", "Outright contradiction or clear scan error. Fix first."),
    ("2 - Important",      "Near-certain error based on format or catalog rules."),
    ("3 - Review",         "Probable error. A human should confirm before applying a fix."),
    ("4 - Advisory",       "Heads-up; likely fine. Glance at it, no action required."),
]


def _rec(text: str, bucket: str) -> dict:
    """Shape each recommendation as {text, bucket} so the UI can label it.
    bucket ∈ {verified, likely, uncertain, summary}."""
    return {"text": text, "bucket": bucket}


def _build_recommendations(by_issue: dict, total_rows: int, flagged_rows: int) -> list[dict]:
    """Generate human-readable, action-oriented recommendations.
    Each recommendation is tagged with the result-bucket it belongs to:
      verified  — near-certain error (came from CRITICAL / HIGH severity flags)
      likely    — probable error (MEDIUM severity)
      uncertain — advisory / weak signal (LOW severity)
      summary   — overall dataset-level framing, not an individual finding
    """
    recs: list[dict] = []
    pct = 100 * flagged_rows / max(total_rows, 1)

    if flagged_rows == 0:
        return [_rec("✓ No issues detected — every row looks clean.", "summary")]

    if pct < 5:
        recs.append(_rec(f"✓ Data quality looks excellent — only {pct:.1f}% of rows need a closer look.", "summary"))
    elif pct < 25:
        recs.append(_rec(f"⚠ About {pct:.1f}% of rows may have problems — start with the ones marked as Verified Matches.", "summary"))
    else:
        recs.append(_rec(f"🚨 About {pct:.1f}% of rows look wrong — this points to a widespread data-entry problem worth fixing at the source.", "summary"))

    # ---- Verified Matches bucket (CRITICAL / HIGH severity issues) ----
    if by_issue.get("imei_identity_contradiction", 0):
        n = by_issue["imei_identity_contradiction"]
        recs.append(_rec(f"🚨 {n} rows share an IMEI number with another row that says it belongs to a different device. An IMEI is globally unique — at least one of each pair is a scan error. Treat as top priority.", "verified"))

    if by_issue.get("same_deal_id_multi_imei", 0):
        n = by_issue["same_deal_id_multi_imei"]
        recs.append(_rec(f"🔗 {n} rows share a Deal ID with another row that has a different IMEI. A Deal ID identifies a single device transaction — two IMEIs under one Deal means at least one IMEI was scanned against the wrong Deal. Pull the physical device(s) for that Deal and keep only the correct IMEI.", "verified"))

    # L18 — BB-truth disagreements (highest-value confirmed errors when BB has overlap)
    bb_brand = by_issue.get("bb_brand_mismatch", 0)
    bb_model = by_issue.get("bb_model_mismatch", 0)
    if bb_brand or bb_model:
        n = bb_brand + bb_model
        recs.append(_rec(f"🧩 {n} devices have a Brand or Model in Master that disagrees with Blackbelt's hardware-tested values. Blackbelt reads identifiers directly from device firmware; treat its values as the truth and update Master.", "verified"))

    if by_issue.get("bb_storage_mismatch", 0):
        n = by_issue["bb_storage_mismatch"]
        recs.append(_rec(f"💾 {n} devices have a different storage size in Master than what Blackbelt physically tested. Storage drives price by 10–30%. Update Master to match BB's tested value.", "verified"))

    if by_issue.get("bb_grade_mismatch", 0):
        n = by_issue["bb_grade_mismatch"]
        recs.append(_rec(f"🏷 {n} devices have a Master grade that disagrees with Blackbelt's automated grade. BB's grading is the test-machine output; use it as the source of truth for pricing.", "verified"))

    if by_issue.get("serial_in_imei_slot", 0):
        n = by_issue["serial_in_imei_slot"]
        recs.append(_rec(f"📝 {n} rows have an alphanumeric serial number written in the IMEI column. Re-scan the IMEI from the device.", "verified"))

    # L19 — Master ↔ Stack reconciliation
    if by_issue.get("master_imei_disagrees_with_stack", 0):
        n = by_issue["master_imei_disagrees_with_stack"]
        recs.append(_rec(f"🔁 {n} devices have a Master IMEI that disagrees with Stack-Bulk's IMEI for the same Deal. Stack records an alternate IMEI — verify which is correct and update Master.", "verified"))

    if by_issue.get("master_stack_grade_mismatch", 0):
        n = by_issue["master_stack_grade_mismatch"]
        recs.append(_rec(f"⚖ {n} devices have a different grade in Master than in Stack-Bulk. Reconcile with the grader and standardise on one source.", "likely"))

    if by_issue.get("master_stack_vat_mismatch", 0):
        n = by_issue["master_stack_vat_mismatch"]
        recs.append(_rec(f"💰 {n} devices have a different VAT classification in Master vs Stack. VAT drives tax handling — reconcile.", "likely"))

    # L21/L22 — BB hardware test failures and refurb parts
    if by_issue.get("bb_test_failed", 0):
        n = by_issue["bb_test_failed"]
        recs.append(_rec(f"🚨 {n} devices have failed Blackbelt hardware tests (battery, screen, sensors, etc.). Do not list as fully functional until repaired and re-tested.", "verified"))

    if by_issue.get("bb_refurbished_parts", 0):
        n = by_issue["bb_refurbished_parts"]
        recs.append(_rec(f"🔧 {n} devices have non-genuine/replaced components per Blackbelt's parts check. Material disclosure for resale — adjust price tier and customer-facing description.", "likely"))

    # L20 — stale inventory
    if by_issue.get("stale_inventory", 0):
        n = by_issue["stale_inventory"]
        recs.append(_rec(f"📅 {n} devices have been in inventory for more than 12 months. Stale inventory ties up capital — review whether to liquidate, return, or keep listing.", "uncertain"))

    if by_issue.get("imei_luhn_fail", 0):
        n = by_issue["imei_luhn_fail"]
        recs.append(_rec(f"🔢 {n} IMEI numbers don't look like real IMEIs (they fail the built-in digit-verification that every genuine IMEI passes). These are usually scanner mis-reads — re-scan the device to confirm.", "verified"))

    if by_issue.get("imei_in_barcode_slot", 0):
        n = by_issue["imei_in_barcode_slot"]
        recs.append(_rec(f"🔄 {n} rows have an IMEI number sitting in the Barcode column — the operator likely scanned the wrong label. Move the value into the IMEI column and re-scan the barcode.", "verified"))

    if by_issue.get("category_model_mismatch", 0):
        n = by_issue["category_model_mismatch"]
        recs.append(_rec(f"📂 {n} rows have a Category that doesn't match the product name (for example, an iPad labelled as a Mobile Phone). Fix the Category field.", "verified"))

    if by_issue.get("tac_cohort_anomaly", 0):
        n = by_issue["tac_cohort_anomaly"]
        recs.append(_rec(f"📡 {n} rows have an IMEI whose first 8 digits (the TAC, which encodes make + model) don't match what every other row of the same model in this batch has. Usually means the operator scanned the IMEI off the wrong phone on the shelf.", "verified"))

    if by_issue.get("model_number_mismatch", 0):
        n = by_issue["model_number_mismatch"]
        recs.append(_rec(f"🔖 {n} rows have a model-number code in the Asset Label (like Apple's A-number or Samsung's SM-code) that doesn't belong to the model named in the label. Either the model name is wrong or the code is wrong.", "verified"))

    if by_issue.get("brand_invalid_value", 0):
        n = by_issue["brand_invalid_value"]
        recs.append(_rec(f"🏷 {n} rows have something that isn't a real brand name in the Brand column (e.g. 'others', 'macbooks'). Replace with the actual manufacturer (Apple, Samsung, etc.).", "verified"))

    if by_issue.get("placeholder_imei", 0):
        recs.append(_rec(f"⚠ Some rows contain fake or test IMEI numbers (like 123456... or all-zeros). Remove these before listing the devices for sale.", "verified"))

    # ---- Likely Matches bucket (MEDIUM severity issues) ----
    if by_issue.get("possible_imei1_imei2_pair", 0):
        n = by_issue["possible_imei1_imei2_pair"]
        recs.append(_rec(f"⚖ {n} rows look like the same phone listed twice — once with each of its two IMEI numbers (dual-SIM phones have IMEI1 and IMEI2). Check whether these are duplicate listings.", "likely"))

    if by_issue.get("qr_code_contradicts_imei", 0):
        n = by_issue["qr_code_contradicts_imei"]
        recs.append(_rec(f"🔀 {n} rows have a valid IMEI in the QR column that's different from the one in the IMEI column. Two different IMEIs can't belong to the same phone — one of them was scanned from a different device.", "likely"))

    if by_issue.get("two_storages_in_label", 0):
        n = by_issue["two_storages_in_label"]
        recs.append(_rec(f"🧩 {n} rows have two different storage sizes in the same Asset Label (e.g. '128GB' and '256GB'). Usually means two listings were merged, or a template wasn't fully edited. Keep only the correct size.", "likely"))

    if by_issue.get("grade_contradicts_damage", 0):
        n = by_issue["grade_contradicts_damage"]
        recs.append(_rec(f"⚖ {n} rows are graded as excellent/good/new but the Asset Label mentions damage (cracked, faulty, no power, etc.). Either the label is from a different unit or the grade was keyed in wrong.", "likely"))

    if by_issue.get("color_not_in_bb_catalog", 0):
        n = by_issue["color_not_in_bb_catalog"]
        recs.append(_rec(f"🎨 {n} rows claim a colour that Blackbelt has never recorded for this model. May be a rare variant — verify the label is correct.", "likely"))

    # ---- Uncertain Matches bucket (LOW severity / advisory) ----
    if by_issue.get("brand_model_not_in_bb_catalog", 0):
        n = by_issue["brand_model_not_in_bb_catalog"]
        recs.append(_rec(f"📚 {n} rows describe a model that Blackbelt has no reference data for. Not an error — but Blackbelt's catalogue should be extended so future scans of this model can be validated automatically.", "uncertain"))

    if by_issue.get("not_in_blackbelt", 0):
        n = by_issue["not_in_blackbelt"]
        pct = 100 * n / max(total_rows, 1)
        recs.append(_rec(f"📡 {n} devices ({pct:.0f}% of inventory) have valid IMEIs but no Blackbelt test on file — coverage gap. Consider including these units (or their models/batches) in the next round of Blackbelt testing so future runs can validate them.", "uncertain"))

    return recs


# ---------------------------------------------------------------------------
# User-friendly Excel export
# ---------------------------------------------------------------------------

from openpyxl.styles import Font, PatternFill, Alignment

_LAYER_NAME = {code: f"{code} – {name}" for code, name, _ in LAYER_INFO}
_LAYER_DESC = {code: desc for code, _, desc in LAYER_INFO}
_PRIORITY_ORDER = {label: idx for idx, (label, _) in enumerate(PRIORITY_EXPLAIN)}

# Column order is fixed by product spec: Deal ID, IMEI, Blackbelt, Stack Bulk,
# Location come first across every download (flagged + clean + ZIP). The
# columns Row #, Priority, Check Type, How to Fix were intentionally removed
# at the manager's request.
_FLAGGED_COLUMNS = [
    "Deal ID", "IMEI", "Blackbelt", "Stack Bulk", "Location",
    "Problem", "Field", "Current Value",
]

_CLEAN_COLUMNS = [
    "Deal ID", "IMEI", "Blackbelt", "Stack Bulk", "Location",
    "Brand", "Asset Label", "Category",
]


def _friendly_flagged(df: pd.DataFrame,
                      bb_by_imei: dict | None = None,
                      stack_by_imei: dict | None = None) -> pd.DataFrame:
    if not len(df):
        return pd.DataFrame(columns=_FLAGGED_COLUMNS)

    bb_by_imei    = bb_by_imei    or {}
    stack_by_imei = stack_by_imei or {}
    issue_info = lambda i, k: ISSUE_INFO.get(i, {}).get(k, "")

    imei_series = df["imei"].astype(str)
    out = pd.DataFrame({
        "Deal ID":           df["appraisal"].astype(str),
        "IMEI":              imei_series,
        "Blackbelt":         imei_series.map(lambda x: bb_by_imei.get(x, "")),
        "Stack Bulk":        imei_series.map(lambda x: stack_by_imei.get(x, "")),
        "Location":          df.get("location_text", pd.Series([""] * len(df))).astype(str),
        "Problem":           df["issue"].map(lambda i: issue_info(i, "problem") or i),
        "Field":             df["field"].astype(str),
        "Current Value":     df["current_value"].astype(str),
    })
    # Sort by severity (CRITICAL first) then row index, so the most important
    # rows are at the top — but those columns are not exported.
    sev_rank = df["severity"].map(SEVERITY_RANK).fillna(99).astype(int).values
    row_idx  = pd.to_numeric(df["co_row"], errors="coerce").fillna(0).astype(int).values
    out = out.assign(_s=sev_rank, _r=row_idx)
    out = out.sort_values(["_s", "_r"]).drop(columns=["_s", "_r"]).reset_index(drop=True)
    return out


def _friendly_clean(df: pd.DataFrame,
                    bb_by_imei: dict | None = None,
                    stack_by_imei: dict | None = None) -> pd.DataFrame:
    if not len(df):
        return pd.DataFrame(columns=_CLEAN_COLUMNS)

    bb_by_imei    = bb_by_imei    or {}
    stack_by_imei = stack_by_imei or {}

    imei_series = df["imei"].astype(str)
    return pd.DataFrame({
        "Deal ID":     df["appraisal"].astype(str) if "appraisal" in df.columns else df.get("asset_id", "").astype(str),
        "IMEI":        imei_series,
        "Blackbelt":   imei_series.map(lambda x: bb_by_imei.get(x, "")),
        "Stack Bulk":  imei_series.map(lambda x: stack_by_imei.get(x, "")),
        "Location":    df.get("location_text", pd.Series([""] * len(df))).astype(str),
        "Brand":       df["brand"].astype(str),
        "Asset Label": df["asset_label"].astype(str),
        "Category":    df["category"].astype(str),
    })


def _style_data_sheet(ws, n_rows: int) -> None:
    """Force text format on every data cell (prevents scientific notation on
    long IMEI strings), bold/colour the header, size columns to fit, freeze
    the header row."""
    header_fill = PatternFill("solid", fgColor="1F2A44")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    if n_rows > 0:
        for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
            for cell in row:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                if cell.value is not None:
                    cell.value = str(cell.value)

    for col_cells in ws.columns:
        max_len = 0
        for c in col_cells:
            v = "" if c.value is None else str(c.value)
            # Cap measured length at 60 so wrapped cells don't blow out the column
            max_len = max(max_len, min(len(v), 60))
        ws.column_dimensions[col_cells[0].column_letter].width = max_len + 3

    ws.freeze_panes = "A2"


def _write_legend_sheet(ws, is_flagged: bool) -> None:
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 90

    header_fill = PatternFill("solid", fgColor="1F2A44")
    header_font = Font(bold=True, color="FFFFFF")
    section_font = Font(bold=True, size=12, color="1F2A44")
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    r = 1
    def write(a, b, *, section=False):
        nonlocal r
        ws.cell(row=r, column=1, value=a).alignment = wrap
        ws.cell(row=r, column=2, value=b).alignment = wrap
        if section:
            ws.cell(row=r, column=1).font = section_font
            ws.cell(row=r, column=2).font = section_font
        r += 1

    if not is_flagged:
        ws.cell(row=r, column=1, value="About this file").font = section_font
        r += 1
        for line in [
            "These rows passed every automated check — no problems detected.",
            "No action required.",
            "",
            "Column meanings:",
        ]:
            write(line, "")
        write("Deal ID",     "Deal / appraisal identifier this row belongs to (Master 'Deal Id' or Stack Bulk 'Appraisal').")
        write("IMEI",        "Device identifier. Stored as text so long numbers display correctly (not as scientific notation).")
        write("Blackbelt",   "What the Blackbelt reference file records for this IMEI (Brand + Model). Blank if no Blackbelt record exists for this IMEI.")
        write("Stack Bulk",  "What the Stack Bulk Upload file records for this IMEI (Asset Label). Blank if Stack Bulk wasn't uploaded or has no row for this IMEI.")
        write("Location",    "Physical warehouse location — Room / Bin / Location for Master Template rows; storage member for Stack Bulk rows.")
        write("Brand",       "Manufacturer.")
        write("Asset Label", "The product description recorded against this unit.")
        write("Category",    "Product category (Mobile Phone, Tablet, Laptop, etc.).")
        return

    # Header row
    ws.cell(row=r, column=1, value="Topic").fill = header_fill
    ws.cell(row=r, column=1).font = header_font
    ws.cell(row=r, column=2, value="Explanation").fill = header_fill
    ws.cell(row=r, column=2).font = header_font
    r += 1

    write("Columns in the 'Flagged Rows' sheet", "", section=True)
    write("Deal ID",           "Deal / appraisal identifier this row belongs to (Master 'Deal Id' or Stack Bulk 'Appraisal').")
    write("IMEI",              "Device identifier (stored as text — long IMEIs are not converted to scientific notation).")
    write("Blackbelt",         "What the Blackbelt reference file records for this IMEI (Brand + Model). Blank if no Blackbelt record exists for this IMEI.")
    write("Stack Bulk",        "What the Stack Bulk Upload file records for this IMEI (Asset Label). Blank if Stack Bulk wasn't uploaded or has no row for this IMEI.")
    write("Location",          "Physical warehouse location — Room / Bin / Location for Master Template rows; storage member for Stack Bulk rows.")
    write("Problem",           "Plain-English description of what looks wrong.")
    write("Field",             "The column in the source data where the problem sits.")
    write("Current Value",     "The value that was found in that column.")
    write("", "")

    write("Check types", "", section=True)
    for code, name, desc in LAYER_INFO:
        write(f"{code} – {name}", desc)
    write("", "")

    write("Problem catalogue", "", section=True)
    seen = set()
    for issue, info in ISSUE_INFO.items():
        key = info["problem"]
        if key in seen:
            continue
        seen.add(key)
        write(info["problem"], info.get("expected", ""))
    write("", "")

    # ------------------------------------------------------------------
    # Device identifiers across categories — clarifies that the IMEI rule
    # only applies to mobile phones and that other categories legitimately
    # carry non-15-digit identifiers in the same column.
    # ------------------------------------------------------------------
    write("About device identifiers (mobile phones + tablets)", "", section=True)
    write("Scope of this audit",
          "Per product spec, the model audits only Mobile Phone and Tablet rows. Smartwatches, laptops, earphones and rows with empty Category are dropped at load time and do not appear in any output below. If you need them included, ask engineering to widen the category filter.")
    write("Mobile phone",
          "Required: 15-digit numeric IMEI that satisfies the Luhn checksum. Anything else — wrong length, non-numeric, or Luhn fail — is flagged as a Confirmed Error. 16-digit values that decompose to a Luhn-valid IMEI body are recognised as IMEISV (see below) and treated as Advisory rather than an error.")
    write("Tablet",
          "Cellular tablets carry a 15-digit IMEI (same rule as phones). Wi-Fi-only tablets carry the manufacturer serial number, which varies in length. The detector does not enforce 15-digit length for tablets — Advisory only — so a non-IMEI serial on a Wi-Fi tablet will not be flagged as a hard error.")
    write("Identifier formats you may see in the column",
          "IMEI = 15 digits, Luhn-valid (mobile phones); IMEISV = 16 digits, IMEI + 2-digit Software Version, no Luhn on the full 16; Manufacturer serial = varies (8–12 typical, alphanumeric for many brands) — common on Wi-Fi tablets.")
    write("Advisory: looks_like_imeisv",
          "When a mobile-phone row has a 16-digit numeric value AND its first 14 digits + a single check digit form a Luhn-valid IMEI, we flag it as Advisory rather than Confirmed Error. This is the IMEISV format — perfectly legitimate, just longer than the canonical 15-digit IMEI. Replace with the 15-digit IMEI for downstream consistency, or keep as IMEISV if your system tracks software version.")
    write("Advisory: not_in_blackbelt",
          "Rows where the IMEI is a valid 15-digit Luhn-passing IMEI but does NOT appear in the Blackbelt reference file. This is a coverage gap — Blackbelt has not tested the device. Use this list to plan the next round of Blackbelt testing. Skipped silently if no Blackbelt file is uploaded or it has no IMEIs.")
    write("", "")

    write("Cross-file reconciliation (L18, L19)", "", section=True)
    write("Trust hierarchy",
          "When the same device appears in multiple files, the data sources rank like this for trust: (1) Blackbelt — values come from a physical hardware test machine, treated as truth; (2) Stack Bulk — sell-side record, second-best; (3) Master Template — manual inventory entries, most prone to keying errors. Disagreements are flagged with the higher-trust source's value as the suggested correction.")
    write("L18 BB ↔ Master",
          "For each Master row whose IMEI matches a Blackbelt IMEI, compare Brand / Model / Storage / Grade / Color / Model Number. Disagreements become Confirmed Errors with BB's value shown as the correct answer. Fires only on rows where BB has a record — invisible until BB coverage is established.")
    write("L19 Master ↔ Stack Bulk",
          "For each Master row whose IMEI matches a Stack Bulk IMEI, cross-check Grade / VAT / Country. When IMEIs don't match but Deal IDs do, surface Stack's IMEI as a likely correction for Master. Fires only when a Stack Bulk file is uploaded.")
    write("L20 Stale inventory",
          "Devices whose Deal date (parsed from the Deal ID) is more than 12 months old. Advisory only — flags long-held inventory worth reviewing for liquidation.")
    write("L21 Blackbelt test failures",
          "When BB recorded one or more failed hardware tests (Battery, Screen, Bluetooth, etc.), surface the device. Do not list as fully functional until repaired and re-tested.")
    write("L22 Blackbelt refurbished parts",
          "When BB's parts check (RP Battery, RP LCD, etc.) flags any component as non-genuine, the device has been previously repaired with non-OEM parts. Material disclosure for resale.")
    write("", "")

    # ------------------------------------------------------------------
    # About the IMEI Luhn check — why we keep this rule, and how it works.
    # Same explainer the team walked through; lives in the report so anyone
    # opening a downloaded file can self-serve.
    # ------------------------------------------------------------------
    write("About the IMEI Luhn check", "", section=True)
    write("Why this rule exists",
          "Every genuine IMEI is required by the GSMA (the global mobile-industry body) to satisfy the Luhn checksum — the same arithmetic check used for credit-card numbers. The standard is in 3GPP TS 23.003 §6.2. Phone manufacturers, carriers, the GSMA IMEI Database, and stolen-phone registries (CEIR, etc.) all enforce it. So if a row in your inventory fails Luhn, the value cannot be a real IMEI — it is mathematically guaranteed to be a scanner mis-read, typo, or fabricated number.")
    write("What it catches",
          "Luhn deterministically catches 100% of single-digit mis-reads and the vast majority of adjacent-digit transpositions. Together, those two error types account for ~95% of all real-world data-entry mistakes — which is why the check sits in the Confirmed Errors bucket: it is not a heuristic, it is a hard rule.")
    write("How the math works",
          "Take the 15 IMEI digits and process them from right to left, indexing 0…14. For each digit at an odd index, double it; if the doubled value exceeds 9, subtract 9. Leave even-index digits as they are. Sum every contribution. The IMEI is valid if and only if that sum is divisible by 10.")
    write("Worked example: 359451189789292 (valid)",
          "Reverse: 2 9 2 9 8 7 9 8 1 1 5 4 9 5 3. Double odd indices and reduce: 2, (9×2=18→9), 2, (9×2=18→9), 8, (7×2=14→5), 9, (8×2=16→7), 1, (1×2=2), 5, (4×2=8), 9, (5×2=10→1), 3. Sum = 2+9+2+9+8+5+9+7+1+2+5+8+9+1+3 = 80. 80 mod 10 = 0  →  VALID.")
    write("Worked example: 866775601760336 (fails)",
          "Same procedure produces a sum of 63. 63 mod 10 = 3, not 0  →  FAILS Luhn. The correct check digit for the body 86677560176033 is 3, so the right IMEI is 866775601760333 — the last digit was mis-read as 6 instead of 3.")
    write("What to do for a failing row",
          "Treat the IMEI as untrusted. Re-scan the device (dial *#06# on phones; for laptops and other categories, read the IMEI label or use the manufacturer's diagnostics). Replace the value in your source system. Do not list the unit until the IMEI is corrected.")


def _write_excel_report(rows: list | pd.DataFrame, out_path: Path, is_flagged: bool,
                        bb_by_imei: dict | None = None,
                        stack_by_imei: dict | None = None) -> None:
    df = pd.DataFrame(rows) if not isinstance(rows, pd.DataFrame) else rows
    if is_flagged:
        friendly = _friendly_flagged(df, bb_by_imei, stack_by_imei)
    else:
        friendly = _friendly_clean(df, bb_by_imei, stack_by_imei)
    sheet_name = "Flagged Rows" if is_flagged else "Clean Rows"

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        friendly.to_excel(xw, sheet_name=sheet_name, index=False)
        _style_data_sheet(xw.sheets[sheet_name], n_rows=len(friendly))
        legend_ws = xw.book.create_sheet("How to Read This")
        _write_legend_sheet(legend_ws, is_flagged=is_flagged)


# ---------------------------------------------------------------------------
# Grade-mismatch focused report
# Surfaces just the bb_grade_mismatch flags as a standalone Excel + summary
# block, since fixing grade mismatches is the highest-priority backend cleanup.
# ---------------------------------------------------------------------------

def _build_grade_mismatch_table(flags_df: pd.DataFrame, co: pd.DataFrame,
                                bb_by_imei: dict) -> pd.DataFrame:
    """One row per bb_grade_mismatch flag with full backend context."""
    cols = ["Deal ID", "AssetId", "IMEI", "Brand", "Asset Label",
            "Backend Grade", "Blackbelt Grade", "Direction", "Suggested Fix"]
    if not len(flags_df):
        return pd.DataFrame(columns=cols)

    g = flags_df[flags_df["issue"] == "bb_grade_mismatch"].copy()
    if not len(g):
        return pd.DataFrame(columns=cols)

    co_idx = co.set_index("co_row")
    pat = re.compile(r"Master='([^']*)', Blackbelt='([^']*)'")

    # Severity ranking for "device is worse than recorded" vs "better than recorded"
    rank = {"a+": 0, "a": 1, "b": 2, "c": 3, "d1": 4, "d2": 5, "f": 6,
            "e": 6, "d": 5}

    def _direction(sk_raw: str, bb_raw: str) -> str:
        sk_n = _grade_normalize(sk_raw)
        bb_n = _grade_normalize(bb_raw)
        if sk_n in rank and bb_n in rank:
            if rank[bb_n] > rank[sk_n]:
                return "Backend optimistic — device is actually worse"
            if rank[bb_n] < rank[sk_n]:
                return "Backend pessimistic — device is actually better"
        return "Different grade"

    rows = []
    for _, f in g.iterrows():
        co_row = co_idx.loc[int(f["co_row"])] if int(f["co_row"]) in co_idx.index else None
        m = pat.search(str(f["current_value"]))
        sk_grade_raw = m.group(1) if m else ""
        bb_grade_raw = m.group(2) if m else ""
        rows.append({
            "Deal ID":         str(f.get("appraisal", "")),
            "AssetId":         str(f.get("asset_id", "")),
            "IMEI":            str(co_row["imei"]) if co_row is not None else "",
            "Brand":           str(co_row["brand"]) if co_row is not None else "",
            "Asset Label":     str(co_row["asset_label"]) if co_row is not None else "",
            "Backend Grade":   sk_grade_raw,
            "Blackbelt Grade": bb_grade_raw,
            "Direction":       _direction(sk_grade_raw, bb_grade_raw),
            "Suggested Fix":   f"Update backend grade to '{bb_grade_raw}' to match Blackbelt's hardware-tested value.",
        })
    return pd.DataFrame(rows, columns=cols)


def _write_grade_mismatch_report(table: pd.DataFrame, out_path: Path) -> None:
    """Write a focused grade-mismatch Excel with a summary tab + detail tab."""
    if not len(table):
        # still write a header-only file so download buttons don't 404
        with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
            table.to_excel(xw, sheet_name="Grade Mismatches", index=False)
            _style_data_sheet(xw.sheets["Grade Mismatches"], n_rows=0)
        return

    # Cross-tab: Backend grade x Blackbelt grade
    cross = (table.groupby(["Backend Grade", "Blackbelt Grade"])
                  .size().reset_index(name="Count")
                  .sort_values("Count", ascending=False))

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        # Detail sheet
        table.to_excel(xw, sheet_name="Grade Mismatches", index=False)
        _style_data_sheet(xw.sheets["Grade Mismatches"], n_rows=len(table))

        # Summary cross-tab sheet
        cross.to_excel(xw, sheet_name="Summary", index=False)
        _style_data_sheet(xw.sheets["Summary"], n_rows=len(cross))

        # Legend
        legend = xw.book.create_sheet("How to Read This")
        legend.column_dimensions["A"].width = 30
        legend.column_dimensions["B"].width = 90
        from openpyxl.styles import Font, PatternFill, Alignment
        wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
        section = Font(bold=True, size=12, color="1F2A44")
        rr = 1
        def w(a, b, *, sec=False):
            nonlocal rr
            legend.cell(row=rr, column=1, value=a).alignment = wrap
            legend.cell(row=rr, column=2, value=b).alignment = wrap
            if sec:
                legend.cell(row=rr, column=1).font = section
                legend.cell(row=rr, column=2).font = section
            rr += 1
        w("About this file", "", sec=True)
        w("Purpose", "Every row where the backend's recorded grade disagrees with "
                    "the grade Blackbelt's hardware test machine produced. "
                    "Treat Blackbelt as authoritative — its grade comes from automated "
                    "diagnostics, not human entry.")
        w("Backend Grade", "What your backend (Stack Bulk or Master Template) currently records.")
        w("Blackbelt Grade", "What Blackbelt's test machine produced for the same IMEI.")
        w("Direction", "Whether the backend is optimistic (overstating condition) or "
                       "pessimistic (understating). Optimistic mismatches are revenue-risk "
                       "(unit listed at higher tier than it should be); pessimistic mismatches "
                       "are pricing-loss (selling for less than the device is worth).")
        w("Summary tab", "Cross-tab of every (Backend, Blackbelt) grade pair so you can spot "
                         "systematic biases (e.g. consistent Stack→D1 vs BB→C drift).")


def _build_bb_by_imei(bb: pd.DataFrame) -> dict:
    """
    Map IMEI (and IMEI2) -> 'Brand Model' string from the Blackbelt file.
    This is the simple display-string lookup used by the report column and
    by L17. The richer record lookup (with all fields) is built separately
    by _build_bb_records.
    """
    out: dict = {}
    for _, r in bb.iterrows():
        label_parts = [str(r.get("brand", "") or "").strip(),
                       str(r.get("model", "") or "").strip()]
        label = " ".join(p for p in label_parts if p).strip()
        if not label:
            continue
        for k in (r.get("imei", ""), r.get("imei2", "")):
            k = str(k or "").strip()
            if k and k not in out:
                out[k] = label
    return out


def _build_bb_secondary_lookups(bb_path: str) -> dict:
    """
    Build IMEI fallback lookups against Blackbelt's secondary identifier
    columns (Serial Number, EID, Custom Value1, DeviceId, MLB serial number).
    Used when the primary IMEI lookup misses — recovers matches when the
    Master IMEI was actually a serial / EID / etc.

    Returns {key_value -> 'Brand Model'} merged across all fallback columns.
    """
    out: dict = {}
    try:
        bb_raw = pd.read_excel(bb_path, sheet_name="Sheet1")
    except Exception:
        return out

    def _label(r):
        b = str(r.get("Manufacturer", "") or "").strip().lower()
        m = str(r.get("Model", "") or "").strip().lower()
        return f"{b} {m}".strip()

    secondary_cols = ("Serial Number", "EID", "Custom Value1",
                      "DeviceId", "MLB serial number", "IMEI2")
    for _, r in bb_raw.iterrows():
        label = _label(r)
        if not label:
            continue
        for col in secondary_cols:
            v = r.get(col)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            k = clean_id(v) or str(v).strip()
            if k and k not in out:
                out[k] = label
    return out


def _build_bb_records(bb_path: str) -> dict:
    """
    Map IMEI (and IMEI2) -> full record of comparable fields from BB.
    Used by L18 to compare Master against BB field-by-field.

    Each record carries: brand, model, model_number, storage_gb, color,
    grade, result, refurbished_parts (list).
    """
    out: dict = {}
    try:
        bb_raw = pd.read_excel(bb_path, sheet_name="Sheet1")
    except Exception:
        return out

    rp_cols = [c for c in bb_raw.columns if c.startswith("RP ") and c.endswith(" OEM Status")]
    test_cols = ["Result", "Battery Test", "Bluetooth", "WiFi", "Speaker",
                 "Microphone", "Screen Test", "Pixel Test", "Camera Front Result",
                 "Camera Rear Result", "Vibration", "GSMA Check"]

    for _, r in bb_raw.iterrows():
        rec = {
            "brand":        brand_canonical(r.get("Manufacturer", "")),
            "model":        norm_text(r.get("Model", "")),
            "model_number": norm_text(r.get("Model Number", "")),
            "storage_gb":   extract_storage_gb(r.get("Handset Memory Size", "")),
            "color":        norm_text(r.get("Device Colour", "")),
            "grade":        norm_text(r.get("Device Grade", "")),
            "result":       norm_text(r.get("Result", "")),
            "tests": {c: str(r.get(c, "") or "").upper() for c in test_cols
                      if pd.notna(r.get(c))},
            "refurbished":  [c.replace("RP ", "").replace(" OEM Status", "")
                             for c in rp_cols
                             if pd.notna(r.get(c))
                             and str(r.get(c)).strip().lower() != "genuine"],
        }
        for k in (clean_id(r.get("IMEI/MEID")), clean_id(r.get("IMEI2"))):
            if k and k not in out:
                out[k] = rec
    return out


def _build_stack_by_imei(stack_path: str) -> dict:
    """Map IMEI -> Asset Label string. Used by the Stack Bulk report column."""
    if not stack_path:
        return {}
    co_stack = _load_company_stackbulk(stack_path,
                                       _detect_company_format(stack_path)[1])
    out: dict = {}
    for _, r in co_stack.iterrows():
        k = str(r.get("imei", "") or "").strip()
        if not k:
            continue
        label = str(r.get("asset_label", "") or "").strip()
        if label and k not in out:
            out[k] = label
    return out


def _build_stack_records(stack_path: str) -> tuple[dict, dict]:
    """
    Return (by_imei, by_deal) full Stack-Bulk record lookups for L19 cross-
    validation. Each record carries: brand, model, asset_label, grade,
    vat_type, country, imei (so we can suggest a corrected IMEI when Master's
    is bad).
    """
    if not stack_path:
        return {}, {}
    co_stack = _load_company_stackbulk(stack_path,
                                       _detect_company_format(stack_path)[1])
    by_imei: dict = {}
    by_deal: dict = {}
    for _, r in co_stack.iterrows():
        rec = {
            "brand":       r.get("brand", ""),
            "asset_label": r.get("asset_label", ""),
            "grade":       r.get("grade", ""),
            "vat_type":    r.get("vat_type", ""),
            "country":     r.get("country", ""),
            "imei":        r.get("imei", ""),
            "imei_shape":  r.get("imei_shape", ""),
        }
        imei = str(r.get("imei", "") or "").strip()
        if imei and imei not in by_imei:
            by_imei[imei] = rec
        deal = str(r.get("appraisal", "") or "").strip()
        if deal and deal not in by_deal:
            by_deal[deal] = rec
    return by_imei, by_deal


def run(bb_path: str = BB_PATH, co_path: str = CO_PATH, out_dir: Path = OUT_DIR,
        stack_path: str | None = None) -> dict:
    out_dir.mkdir(exist_ok=True)

    print("Loading Blackbelt…")
    bb = load_blackbelt(bb_path)
    print(f"  {len(bb)} rows")

    print("Loading Company…")
    co = load_company(co_path)
    print(f"  {len(co)} rows")

    # Per product spec, the model only audits mobile phones and tablets.
    # Other categories (smartwatches, laptops, earphones) are dropped
    # along with rows whose Category is empty.
    KEPT_CATEGORIES = {"mobile phone", "tablet"}
    co_full_count = len(co)
    co = co[co["category"].isin(KEPT_CATEGORIES)].reset_index(drop=True)
    co["co_row"] = co.index
    print(f"  filtered to mobile phone + tablet: {len(co)} rows "
          f"({co_full_count - len(co)} dropped)")

    print("Building catalog…")
    catalog = build_catalog(bb)
    brand_idx = build_brand_idx(catalog)
    print(f"  {len(catalog)} distinct (brand, model) entries")

    # Cross-file lookups built once, used by L17/L18/L19/L21/L22 and the
    # report writer. Multi-key fallback merges BB's primary IMEI lookup with
    # secondary identifier columns (Serial, EID, Custom Value1, etc.) so a
    # Master row whose "IMEI" is actually a serial still resolves.
    bb_by_imei_primary   = _build_bb_by_imei(bb)
    bb_by_imei_secondary = _build_bb_secondary_lookups(bb_path)
    bb_by_imei = {**bb_by_imei_secondary, **bb_by_imei_primary}  # primary wins on collision
    bb_records = _build_bb_records(bb_path)
    stack_by_imei_label = _build_stack_by_imei(stack_path) if stack_path else {}
    stack_by_imei_full, stack_by_deal_full = (_build_stack_records(stack_path)
                                               if stack_path else ({}, {}))

    flags: list[Flag] = []
    for layer_fn, name in [
        (layer1_format,                             "L1  FORMAT"),
        (layer2_scan_slot,                          "L2  SCAN-SLOT"),
        (layer3_intra_row,                          "L3  INTRA-ROW"),
        (lambda df: layer4_catalog(df, brand_idx),  "L4  CATALOG-STORAGE"),
        (layer5_duplicates,                         "L5  DUPLICATES"),
        (layer6_imei1_vs_imei2,                     "L6  IMEI1/2"),
        (layer7_placeholder,                        "L7  PLACEHOLDER"),
        (layer8_brand_validity,                     "L8  BRAND"),
        (layer9_identity_contradiction,             "L9  IDENTITY-CONTRA"),
        (lambda df: layer10_tac_cohort(df, brand_idx),    "L10 TAC-COHORT"),
        (lambda df: layer11_model_number(df, brand_idx),  "L11 MODEL-NUMBER"),
        (lambda df: layer12_color_catalog(df, brand_idx), "L12 COLOR-CATALOG"),
        (layer13_two_storages,                      "L13 TWO-STORAGES"),
        (layer14_grade_damage,                      "L14 GRADE-vs-DAMAGE"),
        (layer15_qr_vs_imei,                        "L15 QR-vs-IMEI"),
        (lambda df: layer16_catalog_gap(df, brand_idx),    "L16 CATALOG-GAP"),
        (lambda df: layer17_blackbelt_coverage(df, bb_by_imei), "L17 BB-COVERAGE"),
        (lambda df: layer18_bb_reconciliation(df, bb_records),  "L18 BB-RECONCILE"),
        (lambda df: layer19_master_stack_recon(df, stack_by_imei_full, stack_by_deal_full),
                                                                "L19 MASTER-STACK"),
        (layer20_stale_inventory,                               "L20 STALE-INVENTORY"),
        (lambda df: layer21_bb_test_failures(df, bb_records),   "L21 BB-TEST-FAIL"),
        (lambda df: layer22_bb_refurbished_parts(df, bb_records), "L22 BB-REFURB"),
    ]:
        added = layer_fn(co)
        print(f"  {name}: {len(added)} flags")
        flags.extend(added)

    # IMEI-recovery enrichment: when L1 flagged a bad IMEI and Stack Bulk has
    # a valid IMEI for the same Deal ID, append the suggested correction to
    # the L1 flag's fix text so the analyst sees the answer inline.
    if stack_by_deal_full:
        deal_to_co = {str(r["appraisal"]): r for _, r in co.iterrows()
                      if r.get("appraisal")}
        bad_imei_issues = {"imei_missing", "imei_luhn_fail", "imei_wrong_length"}
        for f in flags:
            if f.issue not in bad_imei_issues:
                continue
            stack_rec = stack_by_deal_full.get(f.appraisal)
            if not stack_rec:
                continue
            sk_imei = str(stack_rec.get("imei", "") or "").strip()
            sk_shape = stack_rec.get("imei_shape", "")
            if sk_imei and sk_shape == "imei15" and luhn_valid(sk_imei):
                f.suggested_fix = (
                    f.suggested_fix
                    + f"  📡 Stack Bulk has a valid IMEI '{sk_imei}' for the same "
                      f"Deal ID — likely the correct value to use."
                )

    # Sort by severity then row
    flags.sort(key=lambda f: (SEVERITY_RANK[f.severity], f.co_row))

    flags_df = pd.DataFrame([asdict(f) for f in flags])
    flags_df.to_csv(out_dir / "flagged.csv", index=False)

    # Per-row worst flag — advisory-only issues are excluded from severity
    # bucketing so a row whose only flags are coverage/catalog gaps falls
    # through to the clean bucket. They remain in flags_df / by_issue counts
    # so the recommendations sheet still surfaces them.
    real_flags_df = (flags_df[~flags_df["issue"].isin(ADVISORY_ISSUES)]
                     if len(flags_df) else flags_df)
    worst_df = pd.DataFrame()
    if len(real_flags_df):
        worst_df = (real_flags_df.assign(_rank=real_flags_df["severity"].map(SEVERITY_RANK))
                              .sort_values("_rank")
                              .groupby("co_row", as_index=False)
                              .first()
                              .drop(columns="_rank"))
        worst_df.to_csv(out_dir / "per_row.csv", index=False)

    # ------------------------------------------------------------------
    # UI-shaped outputs: bucket each flagged row by its WORST severity.
    # The existing UI labels these as high/medium/low/unmatched cards;
    # we repurpose them as: HIGH=critical-error, MEDIUM=likely-error,
    # LOW=advisory, UNMATCHED=clean-rows (no issues).
    # ------------------------------------------------------------------
    # Enrich worst_df with company-side fields (imei, location_text, brand,
    # asset_label, category) — these aren't on the Flag dataclass but are
    # needed by the new column layout (Deal ID / IMEI / Blackbelt / Stack
    # Bulk / Location).
    if len(worst_df):
        co_extra = co[["co_row", "imei", "location_text", "brand",
                       "asset_label", "category"]].copy()
        co_extra["co_row"] = co_extra["co_row"].astype(int)
        worst_df["co_row"] = worst_df["co_row"].astype(int)
        worst_df = worst_df.merge(co_extra, on="co_row", how="left")

    by_sev = {"CRITICAL": [], "HIGH": [], "MEDIUM": [], "LOW": []}
    if len(worst_df):
        for sev in by_sev:
            sub = worst_df[worst_df["severity"] == sev]
            if len(sub):
                by_sev[sev] = sub.to_dict(orient="records")

    # Combine CRITICAL into HIGH bucket — UI only has 3 severity cards.
    high_rows   = by_sev["CRITICAL"] + by_sev["HIGH"]
    medium_rows = by_sev["MEDIUM"]
    low_rows    = by_sev["LOW"]
    flagged_co_rows = set(int(f.co_row) for f in flags
                          if f.issue not in ADVISORY_ISSUES)
    clean_rows = [
        {"co_row": int(r["co_row"]), "asset_id": str(r["asset_id"]),
         "appraisal": str(r["appraisal"]),
         "imei": r["imei"], "brand": r["brand"],
         "asset_label": r["asset_label"], "category": r["category"],
         "location_text": str(r.get("location_text", "") or "")}
        for _, r in co.iterrows() if int(r["co_row"]) not in flagged_co_rows
    ]

    # bb_by_imei + stack_by_imei_label are built earlier (used by L17/L19).
    # Reused here for the report's Blackbelt and Stack Bulk cross-ref columns.
    stack_by_imei = stack_by_imei_label
    print(f"Cross-ref: {len(bb_by_imei)} BB IMEI keys, {len(stack_by_imei)} Stack Bulk IMEIs")

    # Write the four CSVs the UI download endpoints look for
    pd.DataFrame(high_rows).to_csv(out_dir / UI_FILE_HIGH, index=False)
    pd.DataFrame(medium_rows).to_csv(out_dir / UI_FILE_MEDIUM, index=False)
    pd.DataFrame(low_rows).to_csv(out_dir / UI_FILE_LOW, index=False)
    pd.DataFrame(clean_rows).to_csv(out_dir / UI_FILE_UNMATCHED, index=False)

    # Write user-friendly Excel reports alongside — with plain-English
    # column names, priority/check-type/problem dictionaries, a legend
    # sheet, and long IMEIs stored as text (no scientific notation).
    _write_excel_report(high_rows,   out_dir / XLSX_FILE_HIGH,      is_flagged=True,
                        bb_by_imei=bb_by_imei, stack_by_imei=stack_by_imei)
    _write_excel_report(medium_rows, out_dir / XLSX_FILE_MEDIUM,    is_flagged=True,
                        bb_by_imei=bb_by_imei, stack_by_imei=stack_by_imei)
    _write_excel_report(low_rows,    out_dir / XLSX_FILE_LOW,       is_flagged=True,
                        bb_by_imei=bb_by_imei, stack_by_imei=stack_by_imei)
    _write_excel_report(clean_rows,  out_dir / XLSX_FILE_UNMATCHED, is_flagged=False,
                        bb_by_imei=bb_by_imei, stack_by_imei=stack_by_imei)

    # Dedicated grade-mismatch report — surfaces just bb_grade_mismatch flags
    # in their own file, since fixing grade mismatches is the priority cleanup.
    grade_table = _build_grade_mismatch_table(flags_df, co, bb_by_imei)
    _write_grade_mismatch_report(grade_table, out_dir / XLSX_FILE_GRADES)
    grade_cross = (grade_table.groupby(["Backend Grade", "Blackbelt Grade"])
                              .size().reset_index(name="count")
                              .sort_values("count", ascending=False)
                              .to_dict(orient="records")
                   if len(grade_table) else [])

    # Summary (also fed to the UI as `results`)
    by_severity = dict(Counter(f.severity for f in flags))
    by_layer    = dict(Counter(f.layer for f in flags))
    by_issue    = dict(Counter(f.issue for f in flags))

    n_total    = len(co)
    n_high     = len(high_rows)
    n_medium   = len(medium_rows)
    n_low      = len(low_rows)
    n_clean    = len(clean_rows)

    from datetime import datetime
    summary = {
        # UI-shaped fields (consumed by app.js)
        "total_processed": n_total,
        "total_blackbelt": len(bb),
        "matches": {
            "high_confidence":   {"count": n_high,   "percentage": round(100*n_high/max(n_total,1), 1),
                                  "description": "Confirmed errors — fix immediately"},
            "medium_confidence": {"count": n_medium, "percentage": round(100*n_medium/max(n_total,1), 1),
                                  "description": "Likely errors — verify before fixing"},
            "low_confidence":    {"count": n_low,    "percentage": round(100*n_low/max(n_total,1), 1),
                                  "description": "Advisory — possibly fine, worth a glance"},
            "unmatched":         {"count": n_clean,  "percentage": round(100*n_clean/max(n_total,1), 1),
                                  "description": "No issues detected"},
        },
        "recommendations": _build_recommendations(by_issue, n_total, n_high + n_medium + n_low),
        "processed_at": datetime.now().isoformat(),

        # Grade-mismatch focused block — surfaced as a first-class result so
        # the UI can show it as a dedicated card (priority cleanup target).
        "grade_mismatches": {
            "count":     len(grade_table),
            "matrix":    grade_cross,                # list of {Backend Grade, Blackbelt Grade, count}
            "joinable":  int(sum(1 for _, r in co.iterrows()
                                 if str(r["imei"]) in bb_records)) if bb_records else 0,
        },

        # Detector-internal fields (useful for debugging / exports)
        "detector": {
            "total_flags": len(flags_df),
            "rows_flagged": int(flags_df["co_row"].nunique()) if len(flags_df) else 0,
            "by_severity": by_severity,
            "by_layer": by_layer,
            "by_issue": by_issue,
        },
    }

    (out_dir / UI_FILE_SUMMARY).write_text(json.dumps(summary, indent=2))

    # Console report
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"Company rows:        {n_total}")
    print(f"  HIGH severity:     {n_high}")
    print(f"  MEDIUM severity:   {n_medium}")
    print(f"  LOW severity:      {n_low}")
    print(f"  Clean (no flags):  {n_clean}")
    print(f"Total flags issued:  {len(flags_df)} (rows can have multiple)")
    print("\nBy issue:")
    for k, v in sorted(by_issue.items(), key=lambda x: -x[1]):
        print(f"  {k:<30} {v}")

    print(f"\nArtifacts written to {out_dir.resolve()}:")
    print(f"  - {UI_FILE_HIGH}      ({n_high} rows)")
    print(f"  - {UI_FILE_MEDIUM}    ({n_medium} rows)")
    print(f"  - {UI_FILE_LOW}       ({n_low} rows)")
    print(f"  - {UI_FILE_UNMATCHED} ({n_clean} rows)")
    print(f"  - flagged.csv         (all {len(flags_df)} flags)")
    print(f"  - per_row.csv         (worst flag per row)")
    print(f"  - {UI_FILE_SUMMARY}")
    return summary
    print("  - summary.json    (machine-readable summary)")


if __name__ == "__main__":
    run()
