"""
Blackbelt Mismatch Detection - FastAPI Backend
Production-ready web service with real-time dashboard
"""

import asyncio
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any

from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from mismatch_detector import run as run_detector

# ============================================================================
# Configuration
# ============================================================================

app = FastAPI(title="Blackbelt Mismatch Detection", version="1.0.0")

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create directories
UPLOAD_DIR = Path("uploads")
RESULTS_DIR = Path("results")
UPLOAD_DIR.mkdir(exist_ok=True)
RESULTS_DIR.mkdir(exist_ok=True)

# In-memory job tracking
jobs: Dict[str, Dict[str, Any]] = {}

# ============================================================================
# Models
# ============================================================================

class UploadResponse(BaseModel):
    job_id: str
    status: str
    message: str

class JobStatus(BaseModel):
    job_id: str
    status: str  # pending, processing, completed, failed
    progress: int  # 0-100
    results: Optional[Dict[str, Any]] = None
    error: Optional[str] = None
    created_at: str
    completed_at: Optional[str] = None

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# ============================================================================
# Endpoints
# ============================================================================

@app.get("/")
async def root():
    """Serve the main HTML interface"""
    return FileResponse("static/index.html")

@app.get("/api")
async def api_root():
    """API info endpoint"""
    return {"message": "Blackbelt Mismatch Detection API", "version": "1.0.0"}

@app.post("/api/upload")
async def upload_files(
    blackbelt_file: UploadFile = File(...),
    company_file: UploadFile = File(...),
    stack_file: Optional[UploadFile] = File(None),
    background_tasks: BackgroundTasks = None,
):
    """
    Upload Excel files and start processing.

      blackbelt_file — Blackbelt reference (required)
      company_file   — Stack Bulk Upload, the primary inventory file (required)
      stack_file     — Master Template (optional). Saved with the job for
                       use by upcoming reference/enrichment layers; the
                       current detector run does not consume it yet.

    Returns job_id for tracking.
    """
    try:
        job_id = str(uuid.uuid4())[:8]

        # Save files
        blackbelt_path = UPLOAD_DIR / f"{job_id}_blackbelt.xlsx"
        company_path = UPLOAD_DIR / f"{job_id}_company.xlsx"

        content = await blackbelt_file.read()
        blackbelt_path.write_bytes(content)

        content = await company_file.read()
        company_path.write_bytes(content)

        stack_path = None
        if stack_file is not None:
            stack_path = UPLOAD_DIR / f"{job_id}_stack.xlsx"
            content = await stack_file.read()
            stack_path.write_bytes(content)

        # Initialize job
        jobs[job_id] = {
            "status": "pending",
            "progress": 0,
            "blackbelt_file": str(blackbelt_path),
            "company_file": str(company_path),
            "stack_file": str(stack_path) if stack_path else None,
            "created_at": datetime.now().isoformat(),
            "completed_at": None,
            "results": None,
            "error": None,
        }
        
        # Start background processing
        if background_tasks:
            background_tasks.add_task(process_job, job_id)
        else:
            asyncio.create_task(process_job(job_id))
        
        return UploadResponse(
            job_id=job_id,
            status="pending",
            message="Files uploaded. Processing started.",
        )
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/job/{job_id}")
async def get_job_status(job_id: str):
    """Get current status of a job."""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    return JobStatus(
        job_id=job_id,
        status=job["status"],
        progress=job["progress"],
        results=job.get("results"),
        error=job.get("error"),
        created_at=job["created_at"],
        completed_at=job.get("completed_at"),
    )

@app.get("/api/results/{job_id}")
async def get_results(job_id: str):
    """Get detailed results for a completed job."""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    if job["status"] != "completed":
        raise HTTPException(status_code=400, detail="Job not completed")
    
    return job["results"]

@app.get("/api/download/{job_id}/{report_type}")
async def download_report(job_id: str, report_type: str):
    """Download a report as a user-friendly Excel file."""
    results_dir = RESULTS_DIR / job_id
    
    if not results_dir.exists():
        raise HTTPException(status_code=404, detail="Results not found")

    # Map report types to the Excel files written by the detector
    report_files = {
        # Legacy severity-based downloads (kept for backward compatibility)
        "high":      ("confirmed_errors.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "medium":    ("likely_errors.xlsx",    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "low":       ("advisory_flags.xlsx",   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "unmatched": ("clean_rows.xlsx",       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "summary":   ("summary.json",          "application/json"),
        
        # New category-based downloads
        "brand_mismatch":   ("category_brand_mismatch.xlsx",   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "model_mismatch":   ("category_model_mismatch.xlsx",   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "storage_mismatch": ("category_storage_mismatch.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "grade_mismatch":   ("category_grade_mismatch.xlsx",   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "not_in_blackbelt": ("category_not_in_blackbelt.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }

    # Historical filenames written by the detector — map the nice download
    # name to whatever is actually on disk.
    on_disk = {
        "confirmed_errors.xlsx": "verified_matches.xlsx",
        "likely_errors.xlsx":    "likely_matches.xlsx",
        "advisory_flags.xlsx":   "uncertain_matches.xlsx",
        "clean_rows.xlsx":       "clean_rows.xlsx",
        "summary.json":          "summary.json",
        # Category files use their actual names
        "category_brand_mismatch.xlsx":   "category_brand_mismatch.xlsx",
        "category_model_mismatch.xlsx":   "category_model_mismatch.xlsx",
        "category_storage_mismatch.xlsx": "category_storage_mismatch.xlsx",
        "category_grade_mismatch.xlsx":   "category_grade_mismatch.xlsx",
        "category_not_in_blackbelt.xlsx": "category_not_in_blackbelt.xlsx",
    }

    if report_type not in report_files:
        raise HTTPException(status_code=400, detail="Invalid report type")

    download_name, media_type = report_files[report_type]
    filepath = results_dir / on_disk[download_name]
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Report not found")

    return FileResponse(filepath, media_type=media_type, filename=download_name)

@app.get("/api/export/{job_id}")
async def export_all_results(job_id: str):
    """Download ZIP of all reports."""
    results_dir = RESULTS_DIR / job_id
    
    if not results_dir.exists():
        raise HTTPException(status_code=404, detail="No results found")

    # Bundle the user-facing Excel reports (with the new Deal ID / IMEI /
    # Blackbelt / Stack Bulk / Location layout) plus the JSON summary.
    # Use friendly download names so the contents of the ZIP match what the
    # individual download buttons serve.
    zip_contents = {
        "verified_matches.xlsx":   "confirmed_errors.xlsx",
        "likely_matches.xlsx":     "likely_errors.xlsx",
        "uncertain_matches.xlsx":  "advisory_flags.xlsx",
        "clean_rows.xlsx":         "clean_rows.xlsx",
        "summary.json":            "summary.json",
    }
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for src_name, archive_name in zip_contents.items():
            src = results_dir / src_name
            if src.exists():
                zip_file.write(src, archive_name)

    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=mismatch_results_{job_id}.zip"},
    )

@app.get("/api/download_age/{job_id}/{bucket_type}/{bucket_value}")
async def download_age_bucket(job_id: str, bucket_type: str, bucket_value: str):
    """
    Download detailed Excel file for a specific age bucket.
    
    bucket_type: 'monthly', 'quarterly', 'semi_annual', 'annual', or 'distribution'
    bucket_value: the specific bucket (e.g., '2026-Q1', '2026-04', '0-3mo')
    """
    results_dir = RESULTS_DIR / job_id
    
    if not results_dir.exists():
        raise HTTPException(status_code=404, detail="Results not found")
    
    age_file = results_dir / "product_age.xlsx"
    
    if not age_file.exists():
        raise HTTPException(status_code=404, detail="Age data not found")
    
    import pandas as pd
    import io
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Load the age data
    df = pd.read_excel(age_file, sheet_name="Product Age")
    
    # Filter based on bucket type
    if bucket_type == "distribution":
        # For distribution buckets (0-3mo, 3-6mo, etc.)
        days_ranges = {
            "0-3mo": (0, 90),
            "3-6mo": (91, 180),
            "6-12mo": (181, 365),
            "12+mo": (366, 999999)
        }
        if bucket_value in days_ranges:
            min_days, max_days = days_ranges[bucket_value]
            filtered_df = df[(df["Days Old"] >= min_days) & (df["Days Old"] <= max_days)]
        else:
            raise HTTPException(status_code=400, detail="Invalid distribution bucket")
    else:
        # For time-based buckets (monthly, quarterly, etc.)
        column_map = {
            "monthly": "Monthly",
            "quarterly": "Quarterly",
            "semi_annual": "Semi-annual",
            "annual": "Annual"
        }
        if bucket_type not in column_map:
            raise HTTPException(status_code=400, detail="Invalid bucket type")
        
        column = column_map[bucket_type]
        filtered_df = df[df[column] == bucket_value]
    
    if len(filtered_df) == 0:
        raise HTTPException(status_code=404, detail="No data found for this bucket")
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        filtered_df.to_excel(writer, sheet_name='Devices', index=False)
        
        # Style the worksheet
        workbook = writer.book
        worksheet = writer.sheets['Devices']
        
        # Header styling
        header_fill = PatternFill(start_color="00D4FF", end_color="00D4FF", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # Generate filename
    safe_bucket = bucket_value.replace("/", "-").replace(":", "-")
    filename = f"product_age_{bucket_type}_{safe_bucket}_{job_id}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============================================================================
# Background Task
# ============================================================================

async def process_job(job_id: str):
    """Run the mismatch detector on uploaded files in the background."""
    try:
        job = jobs[job_id]
        job["status"] = "processing"
        job["progress"] = 15
        await asyncio.sleep(0.05)

        output_dir = RESULTS_DIR / job_id
        output_dir.mkdir(exist_ok=True)

        # The detector is CPU-bound and synchronous; run it in a thread so we
        # don't block the event loop and the UI can keep polling progress.
        loop = asyncio.get_event_loop()

        async def bump_progress():
            # Optimistic progress while the detector runs (the layers are fast
            # but the file-load step is the slowest part).
            for pct in (30, 55, 75, 90):
                await asyncio.sleep(0.4)
                if job["status"] == "processing":
                    job["progress"] = pct

        bumper = asyncio.create_task(bump_progress())

        try:
            summary = await loop.run_in_executor(
                None,
                run_detector,
                job["blackbelt_file"],
                job["company_file"],
                output_dir,
                job.get("stack_file"),
            )
        finally:
            bumper.cancel()

        job["results"] = summary
        job["status"] = "completed"
        job["progress"] = 100
        job["completed_at"] = datetime.now().isoformat()

    except Exception as e:
        error_msg = str(e).replace('\n', ' ').strip()
        job["status"] = "failed"
        job["error"] = error_msg
        job["progress"] = 0

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
